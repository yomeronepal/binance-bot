"""
API views for Futures Trading management.
"""
import csv
import io
import json
import logging
from decimal import Decimal
from rest_framework import viewsets, status
from rest_framework.decorators import api_view, permission_classes, action
from rest_framework.permissions import IsAuthenticated, AllowAny, IsAdminUser
from rest_framework.response import Response
from django.db.models import Sum, Count, Q, Avg, Max, Min
from django.db.models.functions import TruncDate
from django.http import HttpResponse
from django.utils import timezone

from signals.models.futures import FuturesTradingSettings, FuturesTrade
from signals.serializers.futures import (
    FuturesTradingSettingsSerializer,
    FuturesTradeSerializer,
    FuturesTradeListSerializer
)

logger = logging.getLogger(__name__)


def _apply_time_filters(queryset, params):
    """
    Filter a FuturesTrade queryset by weekday / hour / month / year on
    ``entry_time``. Weekday and hour are matched in Nepal Time (the
    server-wide convention shared with the public paper-trading
    dashboard); month and year use DB-native lookups against the UTC
    timestamp.

    Each filter is independent and accepts ``'ALL'`` or missing to
    mean "no filter". Invalid values are silently ignored so a bad
    query param can't 500 the dashboard.

    Args:
        queryset: FuturesTrade queryset (or any model with entry_time).
        params: ``request.query_params`` (or a dict-like).

    Returns:
        Filtered queryset.
    """
    from .public_paper_trading import (
        _filter_by_npt_weekday,
        _filter_by_npt_hour,
    )

    weekday = params.get('weekday')
    if weekday and weekday != 'ALL':
        try:
            queryset = _filter_by_npt_weekday(queryset, int(weekday))
        except (ValueError, TypeError):
            pass

    hour = params.get('hour')
    if hour and hour != 'ALL':
        try:
            h = int(hour)
            if 0 <= h <= 23:
                queryset = _filter_by_npt_hour(queryset, h)
        except (ValueError, TypeError):
            pass

    month = params.get('month')
    if month and month != 'ALL':
        try:
            m = int(month)
            if 1 <= m <= 12:
                queryset = queryset.filter(entry_time__month=m)
        except (ValueError, TypeError):
            pass

    year = params.get('year')
    if year and year != 'ALL':
        try:
            queryset = queryset.filter(entry_time__year=int(year))
        except (ValueError, TypeError):
            pass

    return queryset


@api_view(['GET', 'PUT', 'PATCH'])
@permission_classes([IsAdminUser])
def futures_settings(request):
    """
    Get or update futures trading settings.

    GET: Return current settings
    PUT/PATCH: Update settings

    Default settings:
    - trade_amount: $5
    - leverage: 10x
    - max_concurrent_trades: 1
    - is_enabled: False (disabled by default for safety)
    """
    settings_obj = FuturesTradingSettings.get_settings()

    if request.method == 'GET':
        serializer = FuturesTradingSettingsSerializer(settings_obj)
        return Response(serializer.data)

    serializer = FuturesTradingSettingsSerializer(
        settings_obj,
        data=request.data,
        partial=(request.method == 'PATCH')
    )

    if serializer.is_valid():
        serializer.save()
        logger.info(f"Futures settings updated: {request.data}")
        return Response(serializer.data)

    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['POST'])
@permission_classes([IsAdminUser])
def toggle_futures_trading(request):
    """
    Enable or disable futures trading.

    POST body:
    {
        "enabled": true/false
    }
    """
    enabled = request.data.get('enabled')

    if enabled is None:
        return Response(
            {'error': 'enabled field is required'},
            status=status.HTTP_400_BAD_REQUEST
        )

    settings_obj = FuturesTradingSettings.get_settings()
    settings_obj.is_enabled = bool(enabled)
    settings_obj.save()

    action = "enabled" if enabled else "disabled"
    logger.info(f"Futures trading {action}")

    return Response({
        'success': True,
        'message': f'Futures trading {action}',
        'is_enabled': settings_obj.is_enabled
    })


@api_view(['GET'])
@permission_classes([IsAdminUser])
def futures_trades_list(request):
    """
    List all futures trades with optional filtering.

    Query params:
    - status: OPEN, CLOSED_TP, CLOSED_SL, CLOSED_MANUAL, FAILED
    - symbol: Filter by symbol (e.g., BTCUSDT)
    - limit: Number of records (default 50)
    """
    trades = FuturesTrade.objects.select_related('signal').all()
    trades = _apply_time_filters(trades, request.query_params)

    status_filter = request.query_params.get('status')
    if status_filter:
        if status_filter == 'OPEN':
            trades = trades.filter(status='OPEN')
        elif status_filter.startswith('CLOSED'):
            trades = trades.filter(status__startswith='CLOSED')
        else:
            trades = trades.filter(status=status_filter)

    symbol = request.query_params.get('symbol')
    if symbol:
        trades = trades.filter(symbol=symbol)

    priority_filter = request.query_params.get('priority')
    if priority_filter == 'true':
        trades = trades.filter(signal__is_priority=True)
    elif priority_filter == 'false':
        trades = trades.filter(Q(signal__is_priority=False) | Q(signal__isnull=True))

    limit = int(request.query_params.get('limit', 50))
    trades = trades.order_by('-created_at')[:limit]

    serializer = FuturesTradeListSerializer(trades, many=True)
    return Response(serializer.data)


@api_view(['GET'])
@permission_classes([IsAdminUser])
def futures_open_positions(request):
    """Get all open futures positions."""
    open_trades = FuturesTrade.objects.select_related('signal').filter(status='OPEN').order_by('-entry_time')
    serializer = FuturesTradeSerializer(open_trades, many=True)
    return Response(serializer.data)


@api_view(['POST'])
@permission_classes([IsAdminUser])
def close_futures_trade(request, trade_id):
    """
    Manually close a futures trade.

    POST /api/futures/trades/{trade_id}/close/
    """
    try:
        trade = FuturesTrade.objects.get(id=trade_id)
    except FuturesTrade.DoesNotExist:
        return Response(
            {'error': 'Trade not found'},
            status=status.HTTP_404_NOT_FOUND
        )

    if not trade.is_open:
        return Response(
            {'error': f'Trade is already closed (status: {trade.status})'},
            status=status.HTTP_400_BAD_REQUEST
        )

    try:
        from ..services.futures_trader import futures_trading_service

        success = futures_trading_service.close_trade(trade)

        if success:
            trade.refresh_from_db()
            serializer = FuturesTradeSerializer(trade)
            return Response({
                'success': True,
                'message': f'Trade closed successfully',
                'trade': serializer.data
            })
        else:
            return Response(
                {'error': 'Failed to close trade on Binance'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    except Exception as e:
        logger.error(f"Error closing futures trade {trade_id}: {e}")
        return Response(
            {'error': str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['GET'])
@permission_classes([IsAdminUser])
def futures_summary(request):
    """
    Get futures trading summary statistics.

    Returns:
    - Total trades
    - Win rate
    - Total P/L
    - Open positions count
    - Settings status
    """
    settings_obj = FuturesTradingSettings.get_settings()

    all_trades = FuturesTrade.objects.select_related('signal').all()
    all_trades = _apply_time_filters(all_trades, request.query_params)
    closed_trades = all_trades.filter(status__startswith='CLOSED')
    open_trades = all_trades.filter(status='OPEN')

    total_trades = closed_trades.count()
    winning_trades = closed_trades.filter(profit_loss__gt=0).count()
    losing_trades = closed_trades.filter(profit_loss__lt=0).count()

    win_rate = 0
    if total_trades > 0:
        win_rate = (winning_trades / total_trades) * 100

    total_pnl = closed_trades.aggregate(
        total=Sum('profit_loss')
    )['total'] or Decimal('0')

    # Calculate unrealized PnL from open positions
    unrealized_pnl = open_trades.aggregate(
        total=Sum('unrealized_pnl')
    )['total'] or Decimal('0')

    # Get open positions with their live data
    priority_closed = closed_trades.filter(signal__is_priority=True)
    priority_total = priority_closed.count()
    priority_wins = priority_closed.filter(profit_loss__gt=0).count()
    priority_pnl = priority_closed.aggregate(total=Sum('profit_loss'))['total'] or Decimal('0')

    open_positions_data = []
    for trade in open_trades:
        is_priority = getattr(trade.signal, 'is_priority', False) if trade.signal_id else False
        open_positions_data.append({
            'id': trade.id,
            'symbol': trade.symbol,
            'direction': trade.direction,
            'leverage': trade.leverage,
            'entry_price': str(trade.entry_price),
            'mark_price': str(trade.mark_price) if trade.mark_price else None,
            'quantity': str(trade.quantity),
            'position_size_usdt': str(trade.position_size_usdt),
            'unrealized_pnl': str(trade.unrealized_pnl),
            'unrealized_pnl_percentage': str(trade.unrealized_pnl_percentage),
            'liquidation_price': str(trade.liquidation_price) if trade.liquidation_price else None,
            'stop_loss': str(trade.stop_loss),
            'take_profit': str(trade.take_profit),
            'is_priority': is_priority,
            'last_sync_time': trade.last_sync_time.isoformat() if trade.last_sync_time else None,
        })

    return Response({
        'settings': {
            'is_enabled': settings_obj.is_enabled,
            'trade_amount': str(settings_obj.trade_amount),
            'leverage': settings_obj.leverage,
            'effective_position_size': str(settings_obj.effective_position_size),
            'max_concurrent_trades': settings_obj.max_concurrent_trades,
            'allowed_symbols': settings_obj.allowed_symbols,
            'gw_auto_trader_enabled': settings_obj.gw_auto_trader_enabled,
            'total_trading_capital': str(settings_obj.total_trading_capital),
            'max_active_gw_trades': settings_obj.max_active_gw_trades,
        },
        'statistics': {
            'total_trades': total_trades,
            'winning_trades': winning_trades,
            'losing_trades': losing_trades,
            'win_rate': round(win_rate, 2),
            'realized_pnl': str(total_pnl),
            'unrealized_pnl': str(unrealized_pnl),
            'total_pnl': str(total_pnl + unrealized_pnl),
            'open_positions_count': open_trades.count(),
            'priority_trades': priority_total,
            'priority_wins': priority_wins,
            'priority_win_rate': round((priority_wins / priority_total) * 100, 2) if priority_total > 0 else 0,
            'priority_pnl': str(priority_pnl),
        },
        'open_positions': open_positions_data,
    })


@api_view(['GET'])
@permission_classes([IsAdminUser])
def futures_trade_detail(request, trade_id):
    """Get details of a specific futures trade."""
    try:
        trade = FuturesTrade.objects.select_related('signal').get(id=trade_id)
        serializer = FuturesTradeSerializer(trade)
        return Response(serializer.data)
    except FuturesTrade.DoesNotExist:
        return Response(
            {'error': 'Trade not found'},
            status=status.HTTP_404_NOT_FOUND
        )


@api_view(['GET'])
@permission_classes([AllowAny])
def fear_greed_status(request):
    """
    Get current Fear & Greed Index and its effect on trading.

    GET /api/futures/fear-greed/
    """
    from ..services.fear_greed import fetch_fear_greed_index, check_direction_allowed

    settings_obj = FuturesTradingSettings.get_settings()
    fg_data = fetch_fear_greed_index()

    if not fg_data:
        return Response({
            'available': False,
            'enabled': settings_obj.fear_greed_enabled,
            'message': 'Fear & Greed Index temporarily unavailable',
        })

    long_allowed, long_reason = check_direction_allowed(
        'LONG', fg_data['value'],
        settings_obj.fear_greed_short_threshold,
        settings_obj.fear_greed_long_threshold,
    )
    short_allowed, short_reason = check_direction_allowed(
        'SHORT', fg_data['value'],
        settings_obj.fear_greed_short_threshold,
        settings_obj.fear_greed_long_threshold,
    )

    return Response({
        'available': True,
        'enabled': settings_obj.fear_greed_enabled,
        'value': fg_data['value'],
        'classification': fg_data['classification'],
        'source': fg_data.get('source', 'unknown'),
        'components': fg_data.get('components', {}),
        'thresholds': {
            'short_below': settings_obj.fear_greed_short_threshold,
            'long_above': settings_obj.fear_greed_long_threshold,
        },
        'trading_impact': {
            'long_allowed': long_allowed,
            'long_reason': long_reason,
            'short_allowed': short_allowed,
            'short_reason': short_reason,
        },
    })


@api_view(['GET'])
@permission_classes([IsAdminUser])
def futures_report(request):
    """
    Futures trading report with breakdowns.

    GET /api/futures/report/
    """
    base = FuturesTrade.objects.select_related('signal')
    base = _apply_time_filters(base, request.query_params)
    closed = base.filter(status__startswith='CLOSED')
    all_trades = base

    total_closed = closed.count()
    winners = closed.filter(profit_loss__gt=0).count()
    losers = closed.filter(profit_loss__lt=0).count()

    stats = closed.aggregate(
        pnl=Sum('profit_loss'), avg_pnl=Avg('profit_loss'),
        best=Max('profit_loss'), worst=Min('profit_loss'),
    )

    open_trades = all_trades.filter(status='OPEN')
    unrealized = open_trades.aggregate(total=Sum('unrealized_pnl'))['total'] or Decimal('0')

    overall = {
        'total_trades': total_closed,
        'open_trades': open_trades.count(),
        'win_rate': round((winners / total_closed) * 100, 1) if total_closed > 0 else 0,
        'total_pnl': float(stats['pnl'] or 0),
        'unrealized_pnl': float(unrealized),
        'avg_pnl': float(stats['avg_pnl'] or 0),
        'best_trade': float(stats['best'] or 0),
        'worst_trade': float(stats['worst'] or 0),
        'profitable_trades': winners,
        'losing_trades': losers,
    }

    def _agg(qs, field):
        rows = list(qs.values(field).annotate(
            total=Count('id'), wins=Count('id', filter=Q(profit_loss__gt=0)),
            losses=Count('id', filter=Q(profit_loss__lt=0)), pnl=Sum('profit_loss'),
            avg_pnl=Avg('profit_loss'), best=Max('profit_loss'), worst=Min('profit_loss'),
        ).order_by('-pnl'))
        for r in rows:
            r['win_rate'] = round((r['wins'] / r['total']) * 100, 1) if r['total'] > 0 else 0
            for k in ['pnl', 'avg_pnl', 'best', 'worst']:
                r[k] = float(r[k] or 0)
        return rows

    priority_closed = closed.filter(signal__is_priority=True)
    non_priority_closed = closed.filter(Q(signal__is_priority=False) | Q(signal__isnull=True))

    def _prio_stats(qs):
        s = qs.aggregate(total=Count('id'), wins=Count('id', filter=Q(profit_loss__gt=0)), pnl=Sum('profit_loss'))
        t = s['total'] or 0
        return {'total': t, 'wins': s['wins'] or 0,
                'win_rate': round(((s['wins'] or 0) / t) * 100, 1) if t > 0 else 0,
                'pnl': float(s['pnl'] or 0)}

    daily = list(
        closed.filter(exit_time__isnull=False)
        .annotate(day=TruncDate('exit_time'))
        .values('day')
        .annotate(trades=Count('id'), pnl=Sum('profit_loss'), wins=Count('id', filter=Q(profit_loss__gt=0)))
        .order_by('day')
    )
    cum = 0
    for d in daily:
        d['pnl'] = float(d['pnl'] or 0)
        cum += d['pnl']
        d['cumulative_pnl'] = round(cum, 2)
        d['day'] = str(d['day'])

    def _top(qs, asc=False, limit=5):
        order = 'profit_loss' if asc else '-profit_loss'
        filt = Q(profit_loss__lt=0) if asc else Q(profit_loss__gt=0)
        rows = list(qs.filter(filt).order_by(order)[:limit].values(
            'id', 'symbol', 'direction', 'leverage', 'entry_price', 'exit_price',
            'profit_loss', 'profit_loss_percentage'))
        for t in rows:
            for k in ['entry_price', 'exit_price', 'profit_loss', 'profit_loss_percentage']:
                t[k] = float(t[k] or 0)
        return rows

    pnl_list = list(closed.order_by('exit_time').values_list('profit_loss', flat=True))
    max_win = max_lose = streak = 0
    for p in pnl_list:
        if p and p > 0:
            streak = streak + 1 if streak > 0 else 1
            max_win = max(max_win, streak)
        elif p and p < 0:
            streak = streak - 1 if streak < 0 else -1
            max_lose = max(max_lose, abs(streak))

    return Response({
        'overall': overall,
        'by_symbol': _agg(closed, 'symbol'),
        'by_direction': _agg(closed, 'direction'),
        'by_priority': {'priority': _prio_stats(priority_closed), 'non_priority': _prio_stats(non_priority_closed)},
        'daily_pnl': daily,
        'top_winners': _top(closed),
        'top_losers': _top(closed, asc=True),
        'streaks': {'current': streak, 'max_win': max_win, 'max_loss': max_lose},
    })


@api_view(['GET'])
@permission_classes([IsAdminUser])
def futures_balance(request):
    """
    Live Binance futures USDT wallet snapshot.

    Calls /fapi/v2/balance (signed) and returns the USDT row plus a
    ``fetched_at`` timestamp. Admin-only since it reflects real
    account state.

    Returns:
        {
            "balance":          Decimal,     wallet balance
            "available_balance": Decimal,     unencumbered margin
            "unrealized_pnl":    Decimal,     cross-wallet unrealized PnL
            "fetched_at":        ISO timestamp,
            "error":             str | null,
        }
    """
    import asyncio
    from ..services.futures_trader import BinanceFuturesTrader

    async def _go():
        trader = BinanceFuturesTrader(use_testnet=False)
        try:
            return await trader.get_account_balance()
        finally:
            await trader.close()

    payload = {
        'balance': None,
        'available_balance': None,
        'unrealized_pnl': None,
        'fetched_at': timezone.now().isoformat(),
        'error': None,
    }

    try:
        loop = asyncio.new_event_loop()
        try:
            rows = loop.run_until_complete(_go())
        finally:
            loop.close()
    except Exception as exc:
        logger.warning("futures_balance fetch failed: %s", exc)
        payload['error'] = str(exc)
        return Response(payload)

    usdt = next((r for r in (rows or []) if r.get('asset') == 'USDT'), None)
    if not usdt:
        payload['error'] = 'no USDT row in futures balance response'
        return Response(payload)

    def _dec(v):
        try:
            return float(Decimal(str(v)))
        except Exception:
            return None

    payload['balance'] = _dec(usdt.get('balance', '0'))
    payload['available_balance'] = _dec(usdt.get('availableBalance', '0'))
    payload['unrealized_pnl'] = _dec(usdt.get('crossUnPnl', '0'))
    return Response(payload)


FUTURES_EXPORT_COLUMNS = [
    'id', 'symbol', 'direction', 'status', 'leverage', 'margin_type',
    'quantity', 'position_size_usdt',
    'entry_price', 'exit_price', 'stop_loss', 'take_profit',
    'mark_price', 'liquidation_price',
    'profit_loss', 'profit_loss_percentage',
    'unrealized_pnl', 'unrealized_pnl_percentage',
    'max_loss_pct_reached', 'max_profit_pct_reached',
    'cut_loser_triggered', 'current_trailing_tier',
    'binance_order_id', 'binance_exit_order_id',
    'sl_order_id', 'tp_order_id', 'trailing_order_id',
    'signal_id', 'signal_timeframe', 'signal_confidence', 'signal_source',
    'error_message',
    'entry_time', 'exit_time', 'last_sync_time', 'created_at', 'updated_at',
]


def _decimal_or_none(value):
    """Convert Decimal to float; pass through other types."""
    if value is None:
        return None
    if isinstance(value, Decimal):
        return float(value)
    return value


def _iso_or_none(value):
    """Convert datetime to ISO 8601 string."""
    if value is None:
        return None
    return value.isoformat()


def _serialize_trade_for_export(trade):
    """Flatten a FuturesTrade for export to JSON/CSV/Excel."""
    return {
        'id': trade.id,
        'symbol': trade.symbol,
        'direction': trade.direction,
        'status': trade.status,
        'leverage': trade.leverage,
        'margin_type': trade.margin_type,
        'quantity': _decimal_or_none(trade.quantity),
        'position_size_usdt': _decimal_or_none(trade.position_size_usdt),
        'entry_price': _decimal_or_none(trade.entry_price),
        'exit_price': _decimal_or_none(trade.exit_price),
        'stop_loss': _decimal_or_none(trade.stop_loss),
        'take_profit': _decimal_or_none(trade.take_profit),
        'mark_price': _decimal_or_none(trade.mark_price),
        'liquidation_price': _decimal_or_none(trade.liquidation_price),
        'profit_loss': _decimal_or_none(trade.profit_loss),
        'profit_loss_percentage': _decimal_or_none(trade.profit_loss_percentage),
        'unrealized_pnl': _decimal_or_none(trade.unrealized_pnl),
        'unrealized_pnl_percentage': _decimal_or_none(trade.unrealized_pnl_percentage),
        'max_loss_pct_reached': _decimal_or_none(trade.max_loss_pct_reached),
        'max_profit_pct_reached': _decimal_or_none(trade.max_profit_pct_reached),
        'cut_loser_triggered': trade.cut_loser_triggered,
        'current_trailing_tier': trade.current_trailing_tier,
        'binance_order_id': trade.binance_order_id,
        'binance_exit_order_id': trade.binance_exit_order_id,
        'sl_order_id': trade.sl_order_id,
        'tp_order_id': trade.tp_order_id,
        'trailing_order_id': trade.trailing_order_id,
        'signal_id': trade.signal.id if trade.signal else None,
        'signal_timeframe': trade.signal.timeframe if trade.signal else None,
        'signal_confidence': _decimal_or_none(trade.signal.confidence) if trade.signal else None,
        'signal_source': trade.signal.source if trade.signal else None,
        'error_message': trade.error_message or '',
        'entry_time': _iso_or_none(trade.entry_time),
        'exit_time': _iso_or_none(trade.exit_time),
        'last_sync_time': _iso_or_none(trade.last_sync_time),
        'created_at': _iso_or_none(trade.created_at),
        'updated_at': _iso_or_none(trade.updated_at),
    }


def _filtered_export_queryset(request):
    """Build the queryset honoring time filters and optional status/symbol."""
    trades = FuturesTrade.objects.select_related('signal').all()
    trades = _apply_time_filters(trades, request.query_params)

    status_filter = request.query_params.get('status')
    if status_filter:
        if status_filter == 'OPEN':
            trades = trades.filter(status='OPEN')
        elif status_filter.startswith('CLOSED'):
            trades = trades.filter(status__startswith='CLOSED')
        else:
            trades = trades.filter(status=status_filter)

    symbol = request.query_params.get('symbol')
    if symbol:
        trades = trades.filter(symbol=symbol)

    return trades.order_by('-created_at')


def _build_json_export(request, trades, filename):
    """Build a JSON HttpResponse."""
    rows = [_serialize_trade_for_export(t) for t in trades]
    payload = {
        'export_info': {
            'generated_at': timezone.now().isoformat(),
            'generated_by': request.user.username if request.user.is_authenticated else None,
            'total_trades': len(rows),
        },
        'trades': rows,
    }
    response = HttpResponse(
        json.dumps(payload, indent=2),
        content_type='application/json',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def _build_csv_export(trades, filename):
    """Build a CSV HttpResponse."""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    writer = csv.DictWriter(response, fieldnames=FUTURES_EXPORT_COLUMNS, extrasaction='ignore')
    writer.writeheader()
    for trade in trades:
        writer.writerow(_serialize_trade_for_export(trade))
    return response


def _build_excel_export(trades, filename):
    """Build an XLSX HttpResponse, or None if openpyxl is unavailable."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        return None

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Futures Trades'

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1F2937', end_color='1F2937', fill_type='solid')
    sheet.append(FUTURES_EXPORT_COLUMNS)
    for col_idx in range(1, len(FUTURES_EXPORT_COLUMNS) + 1):
        cell = sheet.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill

    for trade in trades:
        row_data = _serialize_trade_for_export(trade)
        sheet.append([row_data.get(col) for col in FUTURES_EXPORT_COLUMNS])

    for col_idx, column_name in enumerate(FUTURES_EXPORT_COLUMNS, start=1):
        letter = sheet.cell(row=1, column=col_idx).column_letter
        sheet.column_dimensions[letter].width = max(12, min(len(column_name) + 4, 30))
    sheet.freeze_panes = 'A2'

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@api_view(['GET'])
@permission_classes([IsAdminUser])
def futures_export(request):
    """
    Export futures trades in JSON, CSV, or XLSX.

    Query params:
        format: ``json`` (default), ``csv``, or ``xlsx`` / ``excel``.
        weekday, hour, month, year: same time filters as the dashboard.
        status: optional trade-status filter.
        symbol: optional symbol filter.

    Returns a file download with a timestamped filename.
    """
    fmt = (request.query_params.get('format') or 'json').lower()
    trades = _filtered_export_queryset(request)
    ts = timezone.now().strftime('%Y%m%d_%H%M%S')

    if fmt == 'csv':
        return _build_csv_export(trades, f'futures_trades_{ts}.csv')

    if fmt in ('xlsx', 'excel'):
        response = _build_excel_export(trades, f'futures_trades_{ts}.xlsx')
        if response is None:
            return Response(
                {'error': 'openpyxl is not installed on the server.'},
                status=status.HTTP_501_NOT_IMPLEMENTED,
            )
        return response

    return _build_json_export(request, trades, f'futures_trades_{ts}.json')
