"""
Signal admin configuration following DRY principles.
"""
import csv
import io
import json
from datetime import datetime, timedelta
from decimal import Decimal
from collections import defaultdict
from django.contrib import admin, messages
from django.utils.html import format_html
from django.utils.safestring import mark_safe
from django.http import HttpResponse
from django.utils import timezone
from .models import Symbol, Signal, UserSubscription, PaperTrade, PaperAccount, TradingSession
from .models.strategy_config import StrategyConfig
from .models.futures import (
    FuturesTradingSettings,
    FuturesTrade,
    FuturesTradeLog,
    BalanceRebalanceLog,
)
from .models.blacklist import BlacklistedSymbol
from .models.push import PushSubscription, NotificationLog
from .models.backtest import BacktestRun, BacktestTrade, BacktestMetric
from .models.top_performers import TopPerformingSymbol


@admin.register(TopPerformingSymbol)
class TopPerformingSymbolAdmin(admin.ModelAdmin):
    """Read-only monthly snapshot — managed by the cron, not by hand."""
    list_display = ('period_start', 'rank', 'symbol', 'total_pnl',
                    'win_rate', 'total_trades', 'calculated_at')
    list_filter = ('period_start',)
    search_fields = ('symbol',)
    ordering = ('-period_start', 'rank')
    readonly_fields = ('symbol', 'period_start', 'period_end', 'rank',
                        'total_trades', 'wins', 'losses', 'win_rate',
                        'total_pnl', 'total_pnl_pct', 'avg_pnl_pct',
                        'best_trade_pct', 'worst_trade_pct', 'calculated_at')

    def has_add_permission(self, request):
        return False  # Created only by the monthly Celery task / backfill.


class BaseModelAdmin(admin.ModelAdmin):
    """
    Base admin class with common configurations (DRY principle).
    """
    readonly_fields = ('created_at', 'updated_at')
    date_hierarchy = 'created_at'
    show_full_result_count = True


@admin.register(Symbol)
class SymbolAdmin(BaseModelAdmin):
    """
    Admin interface for Symbol model.
    """
    list_display = ('symbol', 'exchange', 'active', 'active_signals_count', 'created_at')
    list_filter = ('active', 'exchange', 'created_at')
    search_fields = ('symbol', 'exchange')
    ordering = ('symbol',)
    list_per_page = 50

    actions = ['activate_symbols', 'deactivate_symbols']

    def active_signals_count(self, obj):
        """Display count of active signals for this symbol."""
        count = obj.signals.filter(status='ACTIVE').count()
        if count > 0:
            return format_html('<strong>{}</strong>', count)
        return count
    active_signals_count.short_description = 'Active Signals'

    @admin.action(description='Activate selected symbols')
    def activate_symbols(self, request, queryset):
        """Bulk activate symbols."""
        updated = queryset.update(active=True)
        self.message_user(request, f'{updated} symbols activated successfully.')

    @admin.action(description='Deactivate selected symbols')
    def deactivate_symbols(self, request, queryset):
        """Bulk deactivate symbols."""
        updated = queryset.update(active=False)
        self.message_user(request, f'{updated} symbols deactivated successfully.')


@admin.register(TradingSession)
class TradingSessionAdmin(BaseModelAdmin):
    """
    Admin interface for TradingSession model.
    """
    list_display = (
        'name',
        'session_type_badge',
        'time_window_display',
        'active_days_display',
        'win_rate_display',
        'total_trades_analyzed',
        'auto_generated_badge',
        'active_badge',
        'last_optimized_at',
    )
    list_filter = ('active', 'session_type', 'auto_generated', 'created_at')
    search_fields = ('name', 'description')
    readonly_fields = ('created_at', 'updated_at', 'last_optimized_at')
    ordering = ('start_hour', 'start_minute')
    list_per_page = 50

    fieldsets = (
        ('Basic Information', {
            'fields': ('name', 'session_type', 'description', 'active')
        }),
        ('Time Window (Nepal Time)', {
            'fields': (('start_hour', 'start_minute'), ('end_hour', 'end_minute'))
        }),
        ('Active Days', {
            'fields': ('active_days',),
            'description': 'For GOLDEN_WINDOW sessions, specify active days as a list (0=Monday, 6=Sunday). Leave empty for all days.'
        }),
        ('Optimizer Data', {
            'fields': ('auto_generated', 'win_rate', 'total_trades_analyzed', 'last_optimized_at'),
            'description': 'Auto-populated by the golden window optimizer'
        }),
        ('Timestamps', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        })
    )

    actions = ['activate_sessions', 'deactivate_sessions']

    def session_type_badge(self, obj):
        """Display session type with colored badge."""
        colors = {
            'GOLDEN_WINDOW': '#9333ea',  # purple
            'ACTIVE_TRADING_WINDOW': '#059669',  # green
        }
        color = colors.get(obj.session_type, '#6c757d')
        display_name = dict(obj.SESSION_TYPE_CHOICES).get(obj.session_type, obj.session_type)
        return format_html(
            '<span style="background-color: {}; color: white; padding: 3px 8px; '
            'border-radius: 3px; font-size: 11px;">{}</span>',
            color,
            display_name
        )
    session_type_badge.short_description = 'Type'
    session_type_badge.admin_order_field = 'session_type'

    def time_window_display(self, obj):
        """Display time window in NPT."""
        return f"{obj.start_hour:02d}:{obj.start_minute:02d} - {obj.end_hour:02d}:{obj.end_minute:02d} NPT"
    time_window_display.short_description = 'Time Window'

    def active_days_display(self, obj):
        """Display active days in readable format."""
        if not obj.active_days:
            return format_html('<span style="color: #6c757d;">All Days</span>')
        
        day_names = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']
        active_day_names = [day_names[day] for day in obj.active_days if 0 <= day <= 6]
        return ', '.join(active_day_names) if active_day_names else '-'
    active_days_display.short_description = 'Active Days'

    def win_rate_display(self, obj):
        """Display win rate with color."""
        if obj.win_rate is None:
            return '-'
        wr = float(obj.win_rate)
        color = '#28a745' if wr >= 60 else '#f59e0b' if wr >= 50 else '#dc3545'
        return format_html(
            '<span style="color: {}; font-weight: bold;">{}</span>',
            color, f'{wr:.1f}%'
        )
    win_rate_display.short_description = 'Win Rate'
    win_rate_display.admin_order_field = 'win_rate'

    def auto_generated_badge(self, obj):
        """Display auto-generated status."""
        if obj.auto_generated:
            return format_html(
                '<span style="background-color: #0891b2; color: white; padding: 2px 6px; '
                'border-radius: 3px; font-size: 11px;">AI</span>'
            )
        return format_html(
            '<span style="background-color: #6c757d; color: white; padding: 2px 6px; '
            'border-radius: 3px; font-size: 11px;">Manual</span>'
        )
    auto_generated_badge.short_description = 'Source'
    auto_generated_badge.admin_order_field = 'auto_generated'

    def active_badge(self, obj):
        """Display active status with badge."""
        if obj.active:
            return format_html(
                '<span style="background-color: #28a745; color: white; padding: 3px 8px; '
                'border-radius: 3px; font-size: 11px;">ACTIVE</span>'
            )
        return format_html(
            '<span style="background-color: #6c757d; color: white; padding: 3px 8px; '
            'border-radius: 3px; font-size: 11px;">INACTIVE</span>'
        )
    active_badge.short_description = 'Status'
    active_badge.admin_order_field = 'active'

    @admin.action(description='Activate selected sessions')
    def activate_sessions(self, request, queryset):
        """Bulk activate sessions."""
        updated = queryset.update(active=True)
        self.message_user(request, f'{updated} sessions activated successfully.')

    @admin.action(description='Deactivate selected sessions')
    def deactivate_sessions(self, request, queryset):
        """Bulk deactivate sessions."""
        updated = queryset.update(active=False)
        self.message_user(request, f'{updated} sessions deactivated successfully.')



@admin.register(Signal)
class SignalAdmin(BaseModelAdmin):
    """
    Admin interface for Signal model.
    """
    list_display = (
        'id',
        'priority_badge',
        'symbol_display',
        'direction',
        'market_type',
        'timeframe',
        'entry',
        'sl',
        'tp',
        'confidence_display',
        'rr_ratio',
        'status_badge',
        'created_at'
    )
    list_filter = (
        'is_priority',
        'market_type',
        'direction',
        'status',
        'timeframe',
        'symbol__exchange',
        'created_at',
        'confidence'
    )
    search_fields = (
        'symbol__symbol',
        'description',
        'source',
        'created_by__username'
    )
    ordering = ('-created_at',)
    list_per_page = 25
    date_hierarchy = 'created_at'

    fieldsets = (
        ('Basic Information', {
            'fields': ('symbol', 'direction', 'market_type', 'timeframe', 'status', 'is_priority')
        }),
        ('Price Levels', {
            'fields': ('entry', 'sl', 'tp')
        }),
        ('Signal Details', {
            'fields': ('confidence', 'source', 'description', 'meta')
        }),
        ('Metadata', {
            'fields': ('created_by', 'expires_at', 'created_at', 'updated_at'),
            'classes': ('collapse',)
        })
    )

    autocomplete_fields = ['symbol', 'created_by']
    actions = ['mark_as_executed', 'mark_as_expired', 'mark_as_cancelled']

    def symbol_display(self, obj):
        """Display symbol with exchange."""
        return f"{obj.symbol.symbol} ({obj.symbol.exchange})"
    symbol_display.short_description = 'Symbol'
    symbol_display.admin_order_field = 'symbol__symbol'

    def priority_badge(self, obj):
        """Display priority badge for high win-rate signals."""
        if obj.is_priority:
            return format_html(
                '<span style="background-color: #f59e0b; color: white; padding: 2px 6px; '
                'border-radius: 3px; font-size: 10px; font-weight: bold;">⭐ PRIORITY</span>'
            )
        return ''
    priority_badge.short_description = 'Priority'
    priority_badge.admin_order_field = 'is_priority'

    def confidence_display(self, obj):
        """Display confidence as percentage with color."""
        percentage = int(obj.confidence * 100)
        if obj.confidence >= 0.8:
            color = 'green'
        elif obj.confidence >= 0.5:
            color = 'orange'
        else:
            color = 'red'
        return format_html(
            '<span style="color: {}; font-weight: bold;">{}%</span>',
            color,
            percentage
        )
    confidence_display.short_description = 'Confidence'
    confidence_display.admin_order_field = 'confidence'

    def rr_ratio(self, obj):
        """Display risk/reward ratio."""
        ratio = obj.risk_reward_ratio
        if ratio:
            if ratio >= 2:
                color = 'green'
            elif ratio >= 1:
                color = 'orange'
            else:
                color = 'red'
            return format_html(
                '<span style="color: {};">1:{}</span>',
                color,
                ratio
            )
        return '—'
    rr_ratio.short_description = 'R:R'

    def status_badge(self, obj):
        """Display status with colored badge."""
        colors = {
            'ACTIVE': '#28a745',
            'EXECUTED': '#007bff',
            'EXPIRED': '#6c757d',
            'CANCELLED': '#dc3545'
        }
        color = colors.get(obj.status, '#6c757d')
        return format_html(
            '<span style="background-color: {}; color: white; padding: 3px 8px; '
            'border-radius: 3px; font-size: 11px;">{}</span>',
            color,
            obj.status
        )
    status_badge.short_description = 'Status'
    status_badge.admin_order_field = 'status'

    @admin.action(description='Mark selected signals as EXECUTED')
    def mark_as_executed(self, request, queryset):
        """Bulk mark signals as executed."""
        updated = queryset.update(status='EXECUTED')
        self.message_user(request, f'{updated} signals marked as EXECUTED.')

    @admin.action(description='Mark selected signals as EXPIRED')
    def mark_as_expired(self, request, queryset):
        """Bulk mark signals as expired."""
        updated = queryset.update(status='EXPIRED')
        self.message_user(request, f'{updated} signals marked as EXPIRED.')

    @admin.action(description='Mark selected signals as CANCELLED')
    def mark_as_cancelled(self, request, queryset):
        """Bulk mark signals as cancelled."""
        updated = queryset.update(status='CANCELLED')
        self.message_user(request, f'{updated} signals marked as CANCELLED.')

    def get_queryset(self, request):
        """Optimize queryset with select_related."""
        queryset = super().get_queryset(request)
        return queryset.select_related('symbol', 'created_by')


@admin.register(UserSubscription)
class UserSubscriptionAdmin(BaseModelAdmin):
    """
    Admin interface for UserSubscription model.
    """
    list_display = (
        'user_display',
        'tier_badge',
        'status_badge',
        'is_premium',
        'is_active',
        'expires_at',
        'created_at'
    )
    list_filter = (
        'tier',
        'status',
        'created_at',
        'expires_at'
    )
    search_fields = (
        'user__username',
        'user__email',
        'stripe_customer_id',
        'stripe_subscription_id'
    )
    ordering = ('-created_at',)
    list_per_page = 50

    fieldsets = (
        ('User Information', {
            'fields': ('user',)
        }),
        ('Subscription Details', {
            'fields': ('tier', 'status', 'expires_at')
        }),
        ('Stripe Information', {
            'fields': ('stripe_customer_id', 'stripe_subscription_id'),
            'classes': ('collapse',)
        }),
        ('Timestamps', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        })
    )

    autocomplete_fields = ['user']
    actions = ['upgrade_to_pro', 'upgrade_to_premium', 'downgrade_to_free', 'cancel_subscriptions']

    def user_display(self, obj):
        """Display user with email."""
        return f"{obj.user.username} ({obj.user.email})"
    user_display.short_description = 'User'
    user_display.admin_order_field = 'user__username'

    def tier_badge(self, obj):
        """Display tier with colored badge."""
        colors = {
            'free': '#6c757d',
            'pro': '#007bff',
            'premium': '#ffc107'
        }
        color = colors.get(obj.tier, '#6c757d')
        return format_html(
            '<span style="background-color: {}; color: white; padding: 3px 10px; '
            'border-radius: 3px; font-size: 11px; text-transform: uppercase;">{}</span>',
            color,
            obj.tier
        )
    tier_badge.short_description = 'Tier'
    tier_badge.admin_order_field = 'tier'

    def status_badge(self, obj):
        """Display status with colored badge."""
        colors = {
            'active': '#28a745',
            'inactive': '#6c757d',
            'cancelled': '#dc3545',
            'expired': '#ffc107'
        }
        color = colors.get(obj.status, '#6c757d')
        return format_html(
            '<span style="background-color: {}; color: white; padding: 3px 8px; '
            'border-radius: 3px; font-size: 11px;">{}</span>',
            color,
            obj.status.upper()
        )
    status_badge.short_description = 'Status'
    status_badge.admin_order_field = 'status'

    @admin.action(description='Upgrade to PRO')
    def upgrade_to_pro(self, request, queryset):
        """Bulk upgrade to PRO tier."""
        updated = queryset.update(tier='pro', status='active')
        self.message_user(request, f'{updated} subscriptions upgraded to PRO.')

    @admin.action(description='Upgrade to PREMIUM')
    def upgrade_to_premium(self, request, queryset):
        """Bulk upgrade to PREMIUM tier."""
        updated = queryset.update(tier='premium', status='active')
        self.message_user(request, f'{updated} subscriptions upgraded to PREMIUM.')

    @admin.action(description='Downgrade to FREE')
    def downgrade_to_free(self, request, queryset):
        """Bulk downgrade to FREE tier."""
        updated = queryset.update(tier='free', status='active')
        self.message_user(request, f'{updated} subscriptions downgraded to FREE.')

    @admin.action(description='Cancel subscriptions')
    def cancel_subscriptions(self, request, queryset):
        """Bulk cancel subscriptions."""
        updated = queryset.update(status='cancelled')
        self.message_user(request, f'{updated} subscriptions cancelled.')

    def get_queryset(self, request):
        """Optimize queryset with select_related."""
        queryset = super().get_queryset(request)
        return queryset.select_related('user')


@admin.register(PaperTrade)
class PaperTradeAdmin(BaseModelAdmin):
    """
    Admin interface for PaperTrade model.
    """
    list_display = (
        'id',
        'symbol',
        'direction',
        'market_type',
        'timeframe',
        'confidence_display',
        'entry_price',
        'exit_price',
        'status_badge',
        'profit_loss_display',
        'profit_loss_percentage',
        'user_display',
    )
    list_filter = (
        'status',
        'direction',
        'market_type',
        'timeframe',
        'exit_time'
    )
    search_fields = (
        'symbol',
        'user__username',
        'signal__id'
    )
    # ordering = ('-created_at',)
    list_per_page = 50

    fieldsets = (
        ('Trade Information', {
            'fields': ('user', 'signal', 'symbol', 'direction', 'market_type', 'timeframe', 'confidence', 'status')
        }),
        ('Entry Details', {
            'fields': ('entry_price', 'entry_time', 'position_size', 'quantity', 'leverage')
        }),
        ('Exit Details', {
            'fields': ('stop_loss', 'take_profit', 'exit_price', 'exit_time')
        }),
        ('Performance', {
            'fields': ('profit_loss', 'profit_loss_percentage')
        }),
        ('Timestamps', {
            'fields': ('updated_at',),
            'classes': ('collapse',)
        })
    )

    readonly_fields = ('profit_loss', 'profit_loss_percentage', 'exit_time', 'quantity', 'updated_at')
    autocomplete_fields = ['user', 'signal']
    actions = ['export_to_json', 'export_selected_to_json']

    def user_display(self, obj):
        """Display user."""
        return obj.user.username if obj.user else 'System'
    user_display.short_description = 'User'
    user_display.admin_order_field = 'user__username'

    def confidence_display(self, obj):
        """Display confidence as percentage."""
        if obj.confidence is None:
            return '-'
        pct = obj.confidence * 100
        if pct >= 80:
            color = 'green'
        elif pct >= 70:
            color = 'orange'
        else:
            color = 'gray'
        return format_html(
            '<span style="color: {};">{}%</span>',
            color,
            f'{pct:.0f}'
        )
    confidence_display.short_description = 'Confidence'
    confidence_display.admin_order_field = 'confidence'

    def profit_loss_display(self, obj):
        """Display profit/loss with color."""
        if obj.profit_loss is None:
            return '-'
        pnl = float(obj.profit_loss)
        if pnl > 0:
            color = 'green'
            sign = '+'
        elif pnl < 0:
            color = 'red'
            sign = ''
        else:
            color = 'gray'
            sign = ''
        return format_html(
            '<span style="color: {}; font-weight: bold;">{}{} USDT</span>',
            color,
            sign,
            f'{pnl:.2f}'
        )
    profit_loss_display.short_description = 'P/L'
    profit_loss_display.admin_order_field = 'profit_loss'

    def status_badge(self, obj):
        """Display status with colored badge."""
        colors = {
            'OPEN': '#007bff',
            'CLOSED_TP': '#28a745',
            'CLOSED_SL': '#dc3545',
            'CLOSED_MANUAL': '#6c757d',
            'PENDING': '#ffc107',
            'CANCELLED': '#6c757d'
        }
        color = colors.get(obj.status, '#6c757d')
        return format_html(
            '<span style="background-color: {}; color: white; padding: 3px 8px; '
            'border-radius: 3px; font-size: 11px;">{}</span>',
            color,
            obj.status.replace('_', ' ')
        )
    status_badge.short_description = 'Status'
    status_badge.admin_order_field = 'status'

    def get_queryset(self, request):
        """Optimize queryset with select_related."""
        queryset = super().get_queryset(request)
        return queryset.select_related('user', 'signal')
        
    def get_urls(self):
        from django.urls import path
        urls = super().get_urls()
        my_urls = [
            path('import-json/', self.admin_site.admin_view(self.import_json_view), name='import-json'),
        ]
        return my_urls + urls

    def import_json_view(self, request):
        from django.shortcuts import render, redirect
        from django.contrib import messages
        from django.contrib.auth import get_user_model
        
        User = get_user_model()

        if request.method == "POST":
            json_file = request.FILES.get('json_file')
            if not json_file:
                self.message_user(request, "Please upload a file.", level=messages.ERROR)
                return redirect('..')

            try:
                data = json.load(json_file)
                
                # Handle both structure formats (list of trades OR export dict with 'closed_trades' key)
                trades_to_import = []
                if isinstance(data, list):
                    trades_to_import = data
                elif isinstance(data, dict):
                    # Combine closed and open trades from standard export format
                    trades_to_import.extend(data.get('closed_trades', []))
                    trades_to_import.extend(data.get('open_trades', []))
                
                imported_count = 0
                skipped_count = 0
                
                for trade_data in trades_to_import:
                    try:
                        self._import_trade_from_json(trade_data, User)
                        imported_count += 1
                    except Exception as e:
                        # logger.warning(f"Failed to import trade: {e}")
                        skipped_count += 1
                        
                self.message_user(request, f"Successfully imported {imported_count} trades. Skipped/Failed: {skipped_count}.")
                return redirect('..')
                
            except json.JSONDecodeError:
                self.message_user(request, "Invalid JSON file.", level=messages.ERROR)
                return redirect('..')
            except Exception as e:
                self.message_user(request, f"Error processing file: {str(e)}", level=messages.ERROR)
                return redirect('..')

        context = dict(
           self.admin_site.each_context(request),
           opts=self.model._meta,
        )
        return render(request, "admin/signals/papertrade/import_form.html", context)

    def _import_trade_from_json(self, data, User):
        """Helper to create objects from JSON data"""
        
        # 1. Resolve User
        user = None
        if data.get('username'):
            user = User.objects.filter(username=data['username']).first()
        
        # 2. Resolve or Create Signal
        signal = None
        signal_data = data.get('signal')
        if signal_data:
            # Try to find existing signal by ID first (best effort linkage)
            signal_id = signal_data.get('id')
            if signal_id:
                signal = Signal.objects.filter(id=signal_id).first()
            
            # If not found, try to find by similarity or create new?
            # Creating new signals for historical data might be messy if IDs clash.
            # Strategy: If ID exists, assume it's the correct one. 
            # If not, create a placeholder signal or skip signal linkage if it's just history.
            # Let's try to Create if not exists, but ignore ID collision risks for now by not forcing ID.
            if not signal:
                 # Minimal fields for signal (adjust as needed)
                 # We prioritize linkage over perfect signal recreation
                 pass

        # 3. Create PaperTrade
        # We DO NOT force 'id' from JSON to avoid primary key conflicts with existing data.
        # We rely on creating a NEW record.
        
        # Handle decimal conversions
        def to_dec(val):
            if val is None: return None
            return Decimal(str(val))

        # Handle datetime
        entry_time = None
        if data.get('entry_time'):
            entry_time = datetime.fromisoformat(data['entry_time'].replace('Z', '+00:00'))
            
        exit_time = None
        if data.get('exit_time'):
            exit_time = datetime.fromisoformat(data['exit_time'].replace('Z', '+00:00'))

        gw1 = data.get('is_priority', False)
        gw2 = data.get('is_golden_2', False)
        if entry_time and not gw1:
            from datetime import timedelta as td
            npt = entry_time + td(hours=5, minutes=45)
            dm = npt.hour * 60 + npt.minute
            wd = npt.weekday()
            gw1 = (960 <= dm < 1020) or (1260 <= dm < 1380)
            gw2 = gw1 and (wd in [6, 2, 3])

        PaperTrade.objects.create(
            user=user,
            signal=signal,
            symbol=data.get('symbol'),
            direction=data.get('direction'),
            market_type=data.get('market_type', 'SPOT'),
            timeframe=signal_data.get('timeframe') if signal_data else None,
            confidence=signal_data.get('confidence') if signal_data else None,

            status=data.get('status', 'CLOSED_MANUAL'),

            entry_price=to_dec(data.get('entry_price')),
            entry_time=entry_time,
            position_size=to_dec(data.get('position_size', 100)),
            quantity=to_dec(data.get('quantity')),
            leverage=data.get('leverage'),

            stop_loss=to_dec(data.get('stop_loss')),
            take_profit=to_dec(data.get('take_profit')),

            exit_price=to_dec(data.get('exit_price')),
            exit_time=exit_time,

            profit_loss=to_dec(data.get('profit_loss', 0)),
            profit_loss_percentage=to_dec(data.get('profit_loss_percentage', 0)),

            is_priority=gw1,
            is_golden_2=gw2,
        )

    def _decimal_to_float(self, obj):
        """Helper to convert Decimal and datetime objects for JSON serialization."""
        if obj is None:
            return None
        if isinstance(obj, (Decimal, float, int)):
            return float(obj)
        elif isinstance(obj, datetime):
            return obj.isoformat()
        return obj


    def _calculate_sharpe_ratio(self, trades, risk_free_rate=0.02):
        if not trades or len(trades) < 2:
            return 0
        returns = [t['profit_loss_percentage'] for t in trades if t.get('profit_loss_percentage')]
        if not returns:
            return 0
        avg_return = sum(returns) / len(returns)
        variance = sum((r - avg_return) ** 2 for r in returns) / len(returns)
        std_dev = variance ** 0.5
        if std_dev == 0:
            return 0
        sharpe = (avg_return - risk_free_rate) / std_dev
        return round(sharpe, 3)

    def _calculate_max_drawdown(self, trades):
        if not trades:
            return 0
        cumulative_pnl = 0
        peak = 0
        max_dd = 0
        for trade in trades:
            cumulative_pnl += trade.get('profit_loss', 0)
            if cumulative_pnl > peak:
                peak = cumulative_pnl
            drawdown = peak - cumulative_pnl
            if drawdown > max_dd:
                max_dd = drawdown
        if peak == 0:
            return 0
        return round((max_dd / abs(peak)) * 100, 2)

    def _analyze_consecutive_patterns(self, trades):
        if not trades:
            return {'max_consecutive_wins': 0, 'max_consecutive_losses': 0, 'avg_consecutive_wins': 0, 'avg_consecutive_losses': 0}
        current_streak = 0
        current_type = None
        win_streaks = []
        loss_streaks = []
        for trade in trades:
            is_win = trade.get('profit_loss', 0) > 0
            if current_type is None:
                current_type = 'win' if is_win else 'loss'
                current_streak = 1
            elif (is_win and current_type == 'win') or (not is_win and current_type == 'loss'):
                current_streak += 1
            else:
                if current_type == 'win':
                    win_streaks.append(current_streak)
                else:
                    loss_streaks.append(current_streak)
                current_type = 'win' if is_win else 'loss'
                current_streak = 1
        if current_streak > 0:
            if current_type == 'win':
                win_streaks.append(current_streak)
            else:
                loss_streaks.append(current_streak)
        return {
            'max_consecutive_wins': max(win_streaks) if win_streaks else 0,
            'max_consecutive_losses': max(loss_streaks) if loss_streaks else 0,
            'avg_consecutive_wins': round(sum(win_streaks) / len(win_streaks), 2) if win_streaks else 0,
            'avg_consecutive_losses': round(sum(loss_streaks) / len(loss_streaks), 2) if loss_streaks else 0
        }

    @admin.action(description='Export ALL paper trades to JSON (complete analysis)')
    def export_to_json(self, request, queryset):
        all_trades = PaperTrade.objects.select_related('signal', 'user').all()
        return self._generate_export(request, all_trades, 'paper_trades_complete_export.json')

    @admin.action(description='Export SELECTED trades to JSON')
    def export_selected_to_json(self, request, queryset):
        return self._generate_export(request, queryset, 'paper_trades_selected_export.json')

    def _generate_export(self, request, queryset, filename):
        closed_trades = queryset.filter(status__startswith='CLOSED')
        open_trades = queryset.filter(status__in=['PENDING', 'OPEN'])

        closed_trades_list = []
        for trade in closed_trades:
            trade_data = {
                'id': trade.id, 'symbol': trade.symbol, 'direction': trade.direction,
                'market_type': trade.market_type,
                'entry_price': self._decimal_to_float(trade.entry_price),
                'exit_price': self._decimal_to_float(trade.exit_price),
                'stop_loss': self._decimal_to_float(trade.stop_loss),
                'take_profit': self._decimal_to_float(trade.take_profit),
                'position_size': self._decimal_to_float(trade.position_size),
                'quantity': self._decimal_to_float(trade.quantity), 'leverage': trade.leverage,
                'profit_loss': self._decimal_to_float(trade.profit_loss),
                'profit_loss_percentage': self._decimal_to_float(trade.profit_loss_percentage),
                'status': trade.status,
                'entry_time': trade.entry_time.isoformat() if trade.entry_time else None,
                'exit_time': trade.exit_time.isoformat() if trade.exit_time else None,
                'duration_hours': self._decimal_to_float(trade.duration_hours) if trade.duration_hours else None,
                'risk_reward_ratio': trade.risk_reward_ratio, 'created_at': trade.created_at.isoformat(),
                'timeframe': trade.timeframe, 'confidence': self._decimal_to_float(trade.confidence),
            }
            if trade.signal:
                trade_data['signal'] = {
                    'id': trade.signal.id, 'timeframe': trade.signal.timeframe,
                    'confidence': self._decimal_to_float(trade.signal.confidence),
                    'source': trade.signal.source, 'meta': trade.signal.meta,
                }
            if trade.user:
                trade_data['user_id'] = trade.user.id
                trade_data['username'] = trade.user.username
            closed_trades_list.append(trade_data)

        closed_trades_sorted = sorted(closed_trades_list, key=lambda x: x['exit_time'] if x['exit_time'] else x['created_at'])

        open_trades_list = []
        for trade in open_trades:
            trade_data = {
                'id': trade.id, 'symbol': trade.symbol, 'direction': trade.direction,
                'market_type': trade.market_type,
                'entry_price': self._decimal_to_float(trade.entry_price),
                'stop_loss': self._decimal_to_float(trade.stop_loss),
                'take_profit': self._decimal_to_float(trade.take_profit),
                'position_size': self._decimal_to_float(trade.position_size),
                'quantity': self._decimal_to_float(trade.quantity), 'leverage': trade.leverage,
                'status': trade.status,
                'entry_time': trade.entry_time.isoformat() if trade.entry_time else None,
                'risk_reward_ratio': trade.risk_reward_ratio, 'created_at': trade.created_at.isoformat(),
                'timeframe': trade.timeframe, 'confidence': self._decimal_to_float(trade.confidence),
            }
            if trade.signal:
                trade_data['signal'] = {
                    'id': trade.signal.id, 'timeframe': trade.signal.timeframe,
                    'confidence': self._decimal_to_float(trade.signal.confidence), 'source': trade.signal.source,
                }
            if trade.user:
                trade_data['user_id'] = trade.user.id
                trade_data['username'] = trade.user.username
            open_trades_list.append(trade_data)

        total_closed = len(closed_trades_sorted)
        winning_trades = [t for t in closed_trades_sorted if t['profit_loss'] > 0]
        losing_trades = [t for t in closed_trades_sorted if t['profit_loss'] < 0]
        breakeven_trades = [t for t in closed_trades_sorted if t['profit_loss'] == 0]

        total_profit = sum(t['profit_loss'] for t in closed_trades_sorted)
        total_profit_pct = sum(t['profit_loss_percentage'] for t in closed_trades_sorted)
        win_rate = (len(winning_trades) / total_closed * 100) if total_closed > 0 else 0
        avg_win = sum(t['profit_loss'] for t in winning_trades) / len(winning_trades) if winning_trades else 0
        avg_loss = sum(t['profit_loss'] for t in losing_trades) / len(losing_trades) if losing_trades else 0
        profit_factor = (
            abs(sum(t['profit_loss'] for t in winning_trades)) / abs(sum(t['profit_loss'] for t in losing_trades))
            if losing_trades and sum(t['profit_loss'] for t in losing_trades) != 0 else 0
        )

        sharpe_ratio = self._calculate_sharpe_ratio(closed_trades_sorted)
        max_drawdown = self._calculate_max_drawdown(closed_trades_sorted)
        consecutive_stats = self._analyze_consecutive_patterns(closed_trades_sorted)

        avg_duration = (
            sum(t['duration_hours'] for t in closed_trades_sorted if t['duration_hours']) /
            len([t for t in closed_trades_sorted if t['duration_hours']])
            if any(t['duration_hours'] for t in closed_trades_sorted) else 0
        )

        all_exported_trades = closed_trades_sorted + open_trades_list
        confidences = [t['confidence'] for t in all_exported_trades if t['confidence'] is not None]
        min_trade_confidence = min(confidences) if confidences else 0
        
        timeframe_stats = defaultdict(int)
        for t in all_exported_trades:
            if t['timeframe']:
                timeframe_stats[t['timeframe']] += 1

        by_symbol = defaultdict(lambda: {'total_trades': 0, 'winning_trades': 0, 'losing_trades': 0, 'total_pnl': 0, 'win_rate': 0, 'avg_duration': 0})
        for trade in closed_trades_sorted:
            symbol = trade['symbol']
            by_symbol[symbol]['total_trades'] += 1
            by_symbol[symbol]['total_pnl'] += trade['profit_loss']
            if trade['profit_loss'] > 0:
                by_symbol[symbol]['winning_trades'] += 1
            elif trade['profit_loss'] < 0:
                by_symbol[symbol]['losing_trades'] += 1
            if trade['duration_hours']:
                by_symbol[symbol]['avg_duration'] += trade['duration_hours']

        for symbol, stats in by_symbol.items():
            if stats['total_trades'] > 0:
                stats['win_rate'] = round((stats['winning_trades'] / stats['total_trades']) * 100, 2)
                stats['avg_duration'] = round(stats['avg_duration'] / stats['total_trades'], 2)
            stats['total_pnl'] = round(stats['total_pnl'], 2)

        by_direction = defaultdict(lambda: {'total_trades': 0, 'winning_trades': 0, 'total_pnl': 0, 'win_rate': 0})
        for trade in closed_trades_sorted:
            direction = trade['direction']
            by_direction[direction]['total_trades'] += 1
            by_direction[direction]['total_pnl'] += trade['profit_loss']
            if trade['profit_loss'] > 0:
                by_direction[direction]['winning_trades'] += 1

        for direction, stats in by_direction.items():
            if stats['total_trades'] > 0:
                stats['win_rate'] = round((stats['winning_trades'] / stats['total_trades']) * 100, 2)
            stats['total_pnl'] = round(stats['total_pnl'], 2)

        by_timeframe = defaultdict(lambda: {'total_trades': 0, 'winning_trades': 0, 'total_pnl': 0, 'win_rate': 0})
        for trade in closed_trades_sorted:
            if trade.get('signal'):
                timeframe = trade['signal']['timeframe']
                by_timeframe[timeframe]['total_trades'] += 1
                by_timeframe[timeframe]['total_pnl'] += trade['profit_loss']
                if trade['profit_loss'] > 0:
                    by_timeframe[timeframe]['winning_trades'] += 1

        for timeframe, stats in by_timeframe.items():
            if stats['total_trades'] > 0:
                stats['win_rate'] = round((stats['winning_trades'] / stats['total_trades']) * 100, 2)
            stats['total_pnl'] = round(stats['total_pnl'], 2)

        by_exit_type = defaultdict(lambda: {'count': 0, 'total_pnl': 0, 'avg_pnl': 0})
        for trade in closed_trades_sorted:
            exit_type = trade['status'].replace('CLOSED_', '')
            by_exit_type[exit_type]['count'] += 1
            by_exit_type[exit_type]['total_pnl'] += trade['profit_loss']

        for exit_type, stats in by_exit_type.items():
            if stats['count'] > 0:
                stats['avg_pnl'] = round(stats['total_pnl'] / stats['count'], 2)
            stats['total_pnl'] = round(stats['total_pnl'], 2)

        accounts_data = []
        user_ids = set(t.get('user_id') for t in closed_trades_sorted + open_trades_list if t.get('user_id'))
        for user_id in user_ids:
            try:
                account = PaperAccount.objects.get(user_id=user_id)
                account_data = {
                    'user_id': account.user.id, 'username': account.user.username,
                    'initial_balance': self._decimal_to_float(account.initial_balance),
                    'current_balance': self._decimal_to_float(account.balance),
                    'equity': self._decimal_to_float(account.equity),
                    'total_pnl': self._decimal_to_float(account.total_pnl),
                    'realized_pnl': self._decimal_to_float(account.realized_pnl),
                    'unrealized_pnl': self._decimal_to_float(account.unrealized_pnl),
                    'total_trades': account.total_trades, 'winning_trades': account.winning_trades,
                    'losing_trades': account.losing_trades, 'win_rate': self._decimal_to_float(account.win_rate),
                    'roi_percentage': round((self._decimal_to_float(account.total_pnl) / self._decimal_to_float(account.initial_balance)) * 100, 2) if account.initial_balance > 0 else 0,
                    'auto_trading_enabled': account.auto_trading_enabled,
                    'max_position_size': self._decimal_to_float(account.max_position_size),
                    'max_open_trades': account.max_open_trades,
                    'min_signal_confidence': self._decimal_to_float(account.min_signal_confidence),
                    'open_positions_count': len(account.open_positions),
                    'created_at': account.created_at.isoformat(),
                    'last_trade_at': account.last_trade_at.isoformat() if account.last_trade_at else None,
                }
                accounts_data.append(account_data)
            except PaperAccount.DoesNotExist:
                pass

        time_periods = {'last_7_days': [], 'last_30_days': [], 'last_90_days': [], 'all_time': closed_trades_sorted}
        now = timezone.now()
        for trade in closed_trades_sorted:
            if trade['exit_time']:
                exit_dt = datetime.fromisoformat(trade['exit_time'].replace('Z', '+00:00'))
                days_ago = (now - exit_dt).days
                if days_ago <= 7:
                    time_periods['last_7_days'].append(trade)
                if days_ago <= 30:
                    time_periods['last_30_days'].append(trade)
                if days_ago <= 90:
                    time_periods['last_90_days'].append(trade)

        performance_by_period = {}
        for period_name, trades in time_periods.items():
            if trades:
                wins = [t for t in trades if t['profit_loss'] > 0]
                losses = [t for t in trades if t['profit_loss'] < 0]
                performance_by_period[period_name] = {
                    'total_trades': len(trades), 'winning_trades': len(wins), 'losing_trades': len(losses),
                    'win_rate': round((len(wins) / len(trades)) * 100, 2),
                    'total_pnl': round(sum(t['profit_loss'] for t in trades), 2),
                    'avg_pnl': round(sum(t['profit_loss'] for t in trades) / len(trades), 2),
                    'profit_factor': round(abs(sum(t['profit_loss'] for t in wins)) / abs(sum(t['profit_loss'] for t in losses)), 2) if losses and sum(t['profit_loss'] for t in losses) != 0 else 0,
                    'sharpe_ratio': self._calculate_sharpe_ratio(trades), 'max_drawdown': self._calculate_max_drawdown(trades)
                }
            else:
                performance_by_period[period_name] = {'total_trades': 0, 'winning_trades': 0, 'losing_trades': 0, 'win_rate': 0, 'total_pnl': 0, 'avg_pnl': 0, 'profit_factor': 0, 'sharpe_ratio': 0, 'max_drawdown': 0}

        export_data = {
            'export_info': {
                'generated_at': timezone.now().isoformat(), 'generated_by': request.user.username,
                'total_trades_exported': len(closed_trades_sorted), 'open_trades_exported': len(open_trades_list)
            },
            'summary_statistics': {
                'total_closed_trades': total_closed, 'winning_trades': len(winning_trades),
                'losing_trades': len(losing_trades), 'breakeven_trades': len(breakeven_trades),
                'win_rate': round(win_rate, 2), 'total_profit_loss': round(total_profit, 2),
                'total_profit_loss_percentage': round(total_profit_pct, 2), 'average_win': round(avg_win, 2),
                'average_loss': round(avg_loss, 2), 'profit_factor': round(profit_factor, 2),
                'sharpe_ratio': sharpe_ratio, 'max_drawdown_percentage': max_drawdown,
                'average_duration_hours': round(avg_duration, 2), **consecutive_stats,
                'min_trade_confidence': min_trade_confidence, 'timeframe_summary': dict(timeframe_stats)
            },
            'closed_trades': closed_trades_sorted, 'open_trades': open_trades_list,
            'analysis_by_symbol': dict(by_symbol), 'analysis_by_direction': dict(by_direction),
            'analysis_by_timeframe': dict(by_timeframe), 'analysis_by_exit_type': dict(by_exit_type),
            'performance_by_period': performance_by_period, 'paper_accounts': accounts_data
        }

        response = HttpResponse(json.dumps(export_data, indent=2, default=self._decimal_to_float), content_type='application/json')
        timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
        response['Content-Disposition'] = f'attachment; filename="{filename.replace(".json", "")}_{timestamp}.json"'
        self.message_user(request, f'Successfully exported {total_closed} closed trades and {len(open_trades_list)} open trades. Win Rate: {win_rate:.2f}%, Total P/L: ${total_profit:.2f}')
        return response


@admin.register(PaperAccount)
class PaperAccountAdmin(BaseModelAdmin):
    """
    Admin interface for PaperAccount model (Auto-Trading).
    """
    list_display = (
        'user_display',
        'balance_display',
        'equity_display',
        'total_pnl_display',
        'win_rate_display',
        'total_trades',
        'open_positions_count',
        'auto_trading_status',
        'created_at'
    )
    list_filter = (
        'auto_trading_enabled',
        'auto_trade_spot',
        'auto_trade_futures',
        'created_at'
    )
    search_fields = (
        'user__username',
        'user__email'
    )
    ordering = ('-created_at',)
    list_per_page = 50

    fieldsets = (
        ('User', {
            'fields': ('user',)
        }),
        ('Balance & Equity', {
            'fields': ('initial_balance', 'balance', 'equity')
        }),
        ('Performance Metrics', {
            'fields': (
                'total_pnl',
                'realized_pnl',
                'unrealized_pnl',
                'total_trades',
                'winning_trades',
                'losing_trades',
                'win_rate'
            )
        }),
        ('Risk Management', {
            'fields': ('max_position_size', 'max_open_trades')
        }),
        ('Auto-Trading Settings', {
            'fields': (
                'auto_trading_enabled',
                'auto_trade_spot',
                'auto_trade_futures',
                'min_signal_confidence'
            )
        }),
        ('Open Positions', {
            'fields': ('open_positions',),
            'classes': ('collapse',)
        }),
        ('Timestamps', {
            'fields': ('last_trade_at', 'created_at', 'updated_at'),
            'classes': ('collapse',)
        })
    )

    readonly_fields = (
        'balance',
        'equity',
        'total_pnl',
        'realized_pnl',
        'unrealized_pnl',
        'total_trades',
        'winning_trades',
        'losing_trades',
        'win_rate',
        'open_positions',
        'last_trade_at'
    )
    autocomplete_fields = ['user']
    actions = ['reset_accounts', 'enable_auto_trading', 'disable_auto_trading']

    def user_display(self, obj):
        """Display user."""
        return f"{obj.user.username} ({obj.user.email})"
    user_display.short_description = 'User'
    user_display.admin_order_field = 'user__username'

    def balance_display(self, obj):
        """Display balance."""
        return format_html(
            '<span style="font-weight: bold;">{}</span>',
            f'${float(obj.balance):,.2f}',
        )
    balance_display.short_description = 'Balance'
    balance_display.admin_order_field = 'balance'

    def equity_display(self, obj):
        """Display equity."""
        return format_html(
            '<span style="font-weight: bold;">{}</span>',
            f'${float(obj.equity):,.2f}',
        )
    equity_display.short_description = 'Equity'
    equity_display.admin_order_field = 'equity'

    def total_pnl_display(self, obj):
        """Display total P/L with color."""
        pnl = float(obj.total_pnl)
        if pnl > 0:
            color = 'green'
            sign = '+'
        elif pnl < 0:
            color = 'red'
            sign = ''
        else:
            color = 'gray'
            sign = ''
        return format_html(
            '<span style="color: {}; font-weight: bold;">{}</span>',
            color,
            f'{sign}{pnl:.2f}',
        )
    total_pnl_display.short_description = 'Total P/L'
    total_pnl_display.admin_order_field = 'total_pnl'

    def win_rate_display(self, obj):
        """Display win rate with color."""
        win_rate = float(obj.win_rate)
        if win_rate >= 70:
            color = 'green'
        elif win_rate >= 50:
            color = 'orange'
        else:
            color = 'red'
        return format_html(
            '<span style="color: {}; font-weight: bold;">{}</span>',
            color,
            f'{win_rate:.1f}%',
        )
    win_rate_display.short_description = 'Win Rate'
    win_rate_display.admin_order_field = 'win_rate'

    def open_positions_count(self, obj):
        """Display open positions count."""
        count = len(obj.open_positions)
        if count > 0:
            return format_html('<strong>{}</strong>', count)
        return count
    open_positions_count.short_description = 'Open Positions'

    def auto_trading_status(self, obj):
        """Display auto-trading status."""
        if obj.auto_trading_enabled:
            return format_html(
                '<span style="background-color: #28a745; color: white; padding: 3px 8px; '
                'border-radius: 3px; font-size: 11px;">ENABLED</span>'
            )
        else:
            return format_html(
                '<span style="background-color: #dc3545; color: white; padding: 3px 8px; '
                'border-radius: 3px; font-size: 11px;">DISABLED</span>'
            )
    auto_trading_status.short_description = 'Auto-Trading'
    auto_trading_status.admin_order_field = 'auto_trading_enabled'

    @admin.action(description='Reset accounts to initial state')
    def reset_accounts(self, request, queryset):
        """Bulk reset accounts."""
        for account in queryset:
            account.reset_account()
        self.message_user(request, f'{queryset.count()} accounts reset successfully.')

    @admin.action(description='Enable auto-trading')
    def enable_auto_trading(self, request, queryset):
        """Bulk enable auto-trading."""
        updated = queryset.update(auto_trading_enabled=True)
        self.message_user(request, f'Auto-trading enabled for {updated} accounts.')

    @admin.action(description='Disable auto-trading')
    def disable_auto_trading(self, request, queryset):
        """Bulk disable auto-trading."""
        updated = queryset.update(auto_trading_enabled=False)
        self.message_user(request, f'Auto-trading disabled for {updated} accounts.')

    def get_queryset(self, request):
        """Optimize queryset with select_related."""
        queryset = super().get_queryset(request)
        return queryset.select_related('user')


@admin.register(StrategyConfig)
class StrategyConfigAdmin(admin.ModelAdmin):
    """Admin interface for per-timeframe strategy configuration."""
    list_display = (
        'timeframe', 'is_active_badge',
        'min_confidence', 'rsi_range_display', 'adx_display',
        'sl_tp_display', 'sl_tp_pct_display',
        'updated_at',
    )
    list_filter = ('is_active', 'timeframe')
    list_editable = ('min_confidence',)
    ordering = ('timeframe',)

    fieldsets = (
        ('Timeframe & Status', {
            'fields': ('timeframe', 'is_active'),
        }),
        ('Signal Confidence', {
            'fields': ('min_confidence',),
        }),
        ('RSI Ranges', {
            'fields': (
                'long_rsi_min', 'long_rsi_max',
                'short_rsi_min', 'short_rsi_max',
            ),
            'description': 'RSI thresholds for LONG (oversold) and SHORT (overbought) entries',
        }),
        ('ADX & Volume', {
            'fields': (
                'long_adx_min', 'short_adx_min',
                'long_volume_multiplier', 'short_volume_multiplier',
            ),
        }),
        ('SL/TP - Percentage (Live Signals)', {
            'fields': ('sl_percentage', 'tp_percentage'),
            'description': 'Fixed percentage SL/TP used for live signal generation',
        }),
        ('SL/TP - ATR Multiplier (Backtesting)', {
            'fields': ('sl_atr_multiplier', 'tp_atr_multiplier', 'risk_reward_ratio'),
            'description': 'ATR-based SL/TP used in the backtest engine',
        }),
        ('Indicator Weights', {
            'fields': (
                'macd_weight', 'rsi_weight', 'price_ema_weight',
                'adx_weight', 'ha_weight', 'volume_weight',
                'ema_alignment_weight', 'di_weight', 'bb_weight',
                'volatility_weight', 'supertrend_weight',
                'mfi_weight', 'psar_weight', 'fibonacci_weight',
            ),
            'classes': ('collapse',),
            'description': 'Weights for the 14-indicator confidence scoring system',
        }),
        ('Fibonacci Settings', {
            'fields': (
                'fib_enable_pullback', 'fib_lookback_candles',
                'fib_entry_zone_min', 'fib_entry_zone_max',
            ),
            'classes': ('collapse',),
        }),
        ('Indicator Periods', {
            'fields': (
                'rsi_period', 'atr_period', 'adx_period',
                'macd_fast', 'macd_slow', 'macd_signal',
                'ema_fast', 'ema_medium', 'ema_slow', 'ema_trend',
                'bb_period', 'bb_std_dev', 'volume_ma_period',
                'supertrend_period', 'supertrend_multiplier',
                'mfi_period', 'psar_acceleration', 'psar_maximum',
            ),
            'classes': ('collapse',),
            'description': 'Calculation periods for each indicator (EMA, BB, PSAR, SuperTrend, etc.)',
        }),
        ('Timestamps', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',),
        }),
    )
    readonly_fields = ('created_at', 'updated_at')

    @admin.display(description='Active')
    def is_active_badge(self, obj):
        if obj.is_active:
            return format_html(
                '<span style="background-color: #28a745; color: white; padding: 2px 8px; '
                'border-radius: 3px; font-size: 11px;">ON</span>'
            )
        return format_html(
            '<span style="background-color: #dc3545; color: white; padding: 2px 8px; '
            'border-radius: 3px; font-size: 11px;">OFF</span>'
        )

    @admin.display(description='RSI Range')
    def rsi_range_display(self, obj):
        return format_html(
            'L: <b>{}-{}</b> | S: <b>{}-{}</b>',
            obj.long_rsi_min, obj.long_rsi_max,
            obj.short_rsi_min, obj.short_rsi_max,
        )

    @admin.display(description='ADX Min')
    def adx_display(self, obj):
        return format_html('L: <b>{}</b> | S: <b>{}</b>', obj.long_adx_min, obj.short_adx_min)

    @admin.display(description='SL/TP (ATR)')
    def sl_tp_display(self, obj):
        return format_html(
            'SL: <b>{}x</b> | TP: <b>{}x</b>',
            obj.sl_atr_multiplier, obj.tp_atr_multiplier,
        )

    @admin.display(description='SL/TP (%)')
    def sl_tp_pct_display(self, obj):
        return format_html(
            'SL: <b>{}%</b> | TP: <b>{}%</b>',
            obj.sl_percentage, obj.tp_percentage,
        )


@admin.register(FuturesTradingSettings)
class FuturesTradingSettingsAdmin(admin.ModelAdmin):
    """Admin interface for Futures Trading Settings."""
    list_display = (
        "id", "is_enabled_badge", "trade_amount", "leverage",
        "effective_position_size_display", "max_concurrent_trades",
        "fear_greed_badge", "macro_filter_badge",
        "gw_auto_trader_badge", "dynamic_trailing_badge",
        "cut_loser_badge", "updated_at"
    )
    readonly_fields = ("created_at", "updated_at")

    fieldsets = (
        ('Master Switch', {
            'fields': ('is_enabled',),
            'description': 'Enable or disable all futures trading'
        }),
        ('Capital Management', {
            'fields': (
                'trade_amount',
                'total_trading_capital',
                'max_active_gw_trades',
                'leverage',
                'max_concurrent_trades',
            ),
            'description': 'Configure trading capital and position sizing'
        }),
        ('Signal Filters', {
            'fields': (
                'min_signal_confidence',
                'allowed_symbols',
                'trade_long',
                'trade_short',
            ),
            'description': 'Filter which signals to trade'
        }),
        ('Trading Windows', {
            'fields': (
                'use_trading_window',
                'trade_on_golden_window_2',
                'gw_auto_trader_enabled',
                'daytrade_live_enabled',
            ),
            'description': 'Configure when trading is allowed. daytrade_live_enabled places REAL orders for day-trade signals inside an active Day-Trade Session (still requires the global is_enabled switch).'
        }),
        ('Fear & Greed Filter', {
            'fields': (
                'fear_greed_enabled',
                'fear_greed_short_threshold',
                'fear_greed_long_threshold',
            ),
            'description': (
                'Uses Binance sentiment data (L/S ratio, taker volume, funding rate, OI) '
                'to filter trade direction. F&G <= SHORT threshold: only SHORT. '
                'F&G >= LONG threshold: only LONG. Between: both allowed.'
            )
        }),
        ('Macro Filters (per asset class)', {
            'fields': (
                'crypto_macro_filter_enabled',
                'stock_macro_filter_enabled',
                'commodity_macro_filter_enabled',
                'macro_filter_enabled',
            ),
            'description': (
                "Per-class regime gates at the Binance trade boundary. "
                "Each flag independently controls whether the matching "
                "regime proxy (BTC for crypto, SPY for stocks, XAU for "
                "commodities) can block an order. When OFF, that class "
                "of signal bypasses the gate. Signal-creation tagging "
                "is always on regardless. macro_filter_enabled is the "
                "legacy global flag — kept for backwards compatibility "
                "but no longer consulted at the trade boundary."
            ),
        }),
        ('Neutral Market Reversal', {
            'fields': (
                'neutral_reversal_enabled',
                'neutral_reversal_sl_pct',
                'neutral_reversal_tp_pct',
            ),
            'description': (
                'When F&G is in the neutral zone (between short and long thresholds), '
                'reverse the signal direction and use tight SL/TP. '
                'E.g., LONG signal becomes SHORT with 1.5% SL and 2.5% TP.'
            )
        }),
        ('Cut Loser Strategy', {
            'fields': (
                'cut_loser_enabled',
                'cut_loser_trigger_loss_pct',
                'cut_loser_close_at_pct',
            ),
            'description': 'Close losing trades early when they recover near breakeven'
        }),
        ('Futures Signal-Quality Gate', {
            'fields': (
                'futures_universe_screen_enabled',
                'opposite_exit_enabled',
                'opposite_exit_shadow_mode',
                'opposite_exit_min_confidence',
                'opposite_exit_min_profit_pct',
            ),
            'description': (
                'Universe screening drops illiquid/parabolic symbols before '
                'execution. Opposite-exit arms a drawdown trade when an opposite '
                'day-trade signal appears and closes it once it recovers to '
                'profit (shadow mode logs without acting).'
            )
        }),
        ('Dynamic Trailing Stop', {
            'fields': (
                'dynamic_trailing_enabled',
                'initial_trailing_callback',
                'dynamic_trailing_tiers',
            ),
            'description': 'Trailing stop that tightens as profit grows'
        }),
        ('Timestamps', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )

    def is_enabled_badge(self, obj):
        """Display enabled status as badge."""
        if obj.is_enabled:
            return format_html(
                '<span style="background-color: #28a745; color: white; padding: 3px 10px; '
                'border-radius: 3px; font-weight: bold;">ENABLED</span>'
            )
        return format_html(
            '<span style="background-color: #dc3545; color: white; padding: 3px 10px; '
            'border-radius: 3px; font-weight: bold;">DISABLED</span>'
        )
    is_enabled_badge.short_description = 'Trading'

    def effective_position_size_display(self, obj):
        """Display effective position size."""
        return format_html(
            '<span style="font-weight: bold;">${}</span>',
            f"{float(obj.effective_position_size):.2f}"
        )
    effective_position_size_display.short_description = 'Position Size'

    def fear_greed_badge(self, obj):
        """Display Fear & Greed filter status with live value."""
        if not obj.fear_greed_enabled:
            return format_html(
                '<span style="background-color: #6c757d; color: white; padding: 2px 6px; '
                'border-radius: 3px; font-size: 11px;">F&G OFF</span>'
            )
        try:
            from signals.services.fear_greed import get_fear_greed_value
            val = get_fear_greed_value()
            if val is not None:
                if val <= obj.fear_greed_short_threshold:
                    color = '#dc3545'
                    label = f'F&G {val} SHORT ONLY'
                elif val >= obj.fear_greed_long_threshold:
                    color = '#28a745'
                    label = f'F&G {val} LONG ONLY'
                else:
                    color = '#0d6efd'
                    label = f'F&G {val} BOTH OK'
                return format_html(
                    '<span style="background-color: {}; color: white; padding: 2px 6px; '
                    'border-radius: 3px; font-size: 11px;">{}</span>',
                    color, label
                )
        except Exception:
            pass
        return format_html(
            '<span style="background-color: #f59e0b; color: white; padding: 2px 6px; '
            'border-radius: 3px; font-size: 11px;">F&G ON</span>'
        )
    fear_greed_badge.short_description = 'F&G'

    def gw_auto_trader_badge(self, obj):
        """Display golden window auto trader status."""
        if obj.gw_auto_trader_enabled:
            return format_html(
                '<span style="background-color: #9333ea; color: white; padding: 2px 6px; '
                'border-radius: 3px; font-size: 11px;">GW AUTO</span>'
            )
        return format_html(
            '<span style="background-color: #6c757d; color: white; padding: 2px 6px; '
            'border-radius: 3px; font-size: 11px;">GW OFF</span>'
        )
    gw_auto_trader_badge.short_description = 'GW Auto'

    def dynamic_trailing_badge(self, obj):
        """Display dynamic trailing stop status."""
        if obj.dynamic_trailing_enabled:
            return format_html(
                '<span style="background-color: #059669; color: white; padding: 2px 6px; '
                'border-radius: 3px; font-size: 11px;">TRAILING</span>'
            )
        return format_html(
            '<span style="background-color: #6c757d; color: white; padding: 2px 6px; '
            'border-radius: 3px; font-size: 11px;">FIXED SL</span>'
        )
    dynamic_trailing_badge.short_description = 'SL Type'

    def macro_filter_badge(self, obj):
        """Display BTC macro filter status."""
        if getattr(obj, 'macro_filter_enabled', False):
            return format_html(
                '<span style="background-color: #10b981; color: white; padding: 2px 6px; '
                'border-radius: 3px; font-size: 11px;">MACRO ON</span>'
            )
        return format_html(
            '<span style="background-color: #6c757d; color: white; padding: 2px 6px; '
            'border-radius: 3px; font-size: 11px;">MACRO OFF</span>'
        )
    macro_filter_badge.short_description = 'BTC Macro'

    def cut_loser_badge(self, obj):
        """Display cut loser status."""
        if obj.cut_loser_enabled:
            return format_html(
                '<span style="background-color: #f59e0b; color: white; padding: 2px 6px; '
                'border-radius: 3px; font-size: 11px;">CUT LOSER</span>'
            )
        return format_html(
            '<span style="background-color: #6c757d; color: white; padding: 2px 6px; '
            'border-radius: 3px; font-size: 11px;">STANDARD</span>'
        )
    cut_loser_badge.short_description = 'Risk Mgmt'

    def has_add_permission(self, request):
        return not FuturesTradingSettings.objects.exists()

    def has_delete_permission(self, request, obj=None):
        return False


@admin.register(FuturesTrade)
class FuturesTradeAdmin(BaseModelAdmin):
    """Admin interface for Futures Trades."""
    change_list_template = "admin/signals/futurestrade/change_list.html"
    actions = [
        'export_selected_to_json',
        'export_selected_to_csv',
        'export_selected_to_excel',
    ]
    list_display = (
        "id", "symbol", "direction_badge", "leverage",
        "entry_price", "mark_price", "status_badge",
        "unrealized_pnl_display", "pnl_display",
        "trailing_tier_badge", "cut_loser_badge",
        "entry_time"
    )
    list_filter = ("status", "direction", "symbol", "cut_loser_triggered", "entry_time")
    search_fields = ("symbol", "binance_order_id")
    ordering = ("-created_at",)
    readonly_fields = (
        "signal", "binance_order_id", "binance_exit_order_id", "trailing_order_id",
        "entry_time", "exit_time", "profit_loss", "profit_loss_percentage",
        "mark_price", "unrealized_pnl", "unrealized_pnl_percentage",
        "liquidation_price", "margin_type", "last_sync_time",
        "cut_loser_triggered", "max_loss_pct_reached", "max_profit_pct_reached",
        "current_trailing_tier",
        "created_at", "updated_at"
    )

    fieldsets = (
        ('Trade Information', {
            'fields': ('signal', 'symbol', 'direction', 'status')
        }),
        ('Position Details', {
            'fields': ('leverage', 'quantity', 'position_size_usdt', 'margin_type')
        }),
        ('Price Levels', {
            'fields': ('entry_price', 'stop_loss', 'take_profit', 'exit_price')
        }),
        ('Live Data (from Binance)', {
            'fields': (
                'mark_price', 'liquidation_price',
                'unrealized_pnl', 'unrealized_pnl_percentage',
                'last_sync_time'
            ),
            'description': 'Real-time data synced from Binance'
        }),
        ('Realized Performance', {
            'fields': ('profit_loss', 'profit_loss_percentage')
        }),
        ('Cut Loser Tracking', {
            'fields': (
                'cut_loser_triggered',
                'max_loss_pct_reached',
            ),
            'description': 'Tracking for cut-loser risk management'
        }),
        ('Dynamic Trailing Stop Tracking', {
            'fields': (
                'current_trailing_tier',
                'max_profit_pct_reached',
                'trailing_order_id',
            ),
            'description': 'Tracking for dynamic trailing stop'
        }),
        ('Binance Order IDs', {
            'fields': ('binance_order_id', 'binance_exit_order_id', 'error_message'),
            'classes': ('collapse',)
        }),
        ('Timestamps', {
            'fields': ('entry_time', 'exit_time', 'created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )

    @admin.display(description='Direction')
    def direction_badge(self, obj):
        """Display direction as colored badge."""
        color = "#28a745" if obj.direction == "LONG" else "#dc3545"
        return format_html(
            '<span style="background-color: {}; color: white; padding: 3px 8px; '
            'border-radius: 3px; font-weight: bold;">{}</span>',
            color, obj.direction
        )

    @admin.display(description='Status')
    def status_badge(self, obj):
        """Display status with colored badge."""
        colors = {
            'PENDING': '#ffc107',
            'OPEN': '#007bff',
            'CLOSED_TP': '#28a745',
            'CLOSED_SL': '#dc3545',
            'CLOSED_MANUAL': '#6c757d',
            'FAILED': '#dc3545',
            'CANCELLED': '#6c757d'
        }
        color = colors.get(obj.status, '#6c757d')
        return format_html(
            '<span style="background-color: {}; color: white; padding: 3px 8px; '
            'border-radius: 3px; font-size: 11px;">{}</span>',
            color, obj.status.replace('_', ' ')
        )

    @admin.display(description='Realized P/L')
    def pnl_display(self, obj):
        """Display realized P/L with color."""
        if obj.profit_loss is None:
            return '-'
        pnl = float(obj.profit_loss or 0)
        if pnl > 0:
            color = 'green'
            sign = '+'
        elif pnl < 0:
            color = 'red'
            sign = ''
        else:
            color = 'gray'
            sign = ''
        return format_html(
            '<span style="color: {}; font-weight: bold;">{} USDT</span>',
            color, f"{sign}{pnl:.4f}"
        )

    @admin.display(description='Unrealized')
    def unrealized_pnl_display(self, obj):
        """Display unrealized P/L with color."""
        if obj.unrealized_pnl is None or obj.status != 'OPEN':
            return '-'
        pnl = float(obj.unrealized_pnl or 0)
        pnl_pct = float(obj.unrealized_pnl_percentage or 0)
        if pnl > 0:
            color = 'green'
            sign = '+'
        elif pnl < 0:
            color = 'red'
            sign = ''
        else:
            color = 'gray'
            sign = ''
        return format_html(
            '<span style="color: {}; font-weight: bold;">{} ({}%)</span>',
            color, f"{sign}{pnl:.2f}", f"{sign}{pnl_pct:.2f}"
        )

    @admin.display(description='Trail Tier')
    def trailing_tier_badge(self, obj):
        """Display current trailing tier."""
        tier = obj.current_trailing_tier or 0
        if tier == 0:
            return format_html(
                '<span style="background-color: #6c757d; color: white; padding: 2px 6px; '
                'border-radius: 3px; font-size: 11px;">T0</span>'
            )
        color = '#059669' if tier >= 2 else '#10b981'
        return format_html(
            '<span style="background-color: {}; color: white; padding: 2px 6px; '
            'border-radius: 3px; font-size: 11px;">T{}</span>',
            color, tier
        )

    @admin.display(description='Cut Loser')
    def cut_loser_badge(self, obj):
        """Display cut loser status."""
        if obj.cut_loser_triggered:
            return format_html(
                '<span style="background-color: #f59e0b; color: white; padding: 2px 6px; '
                'border-radius: 3px; font-size: 11px;">TRIGGERED</span>'
            )
        return '-'

    def get_queryset(self, request):
        """Optimize queryset."""
        return super().get_queryset(request).select_related('signal')

    EXPORT_COLUMNS = [
        'id', 'symbol', 'direction', 'status', 'leverage', 'margin_type',
        'quantity', 'position_size_usdt',
        'entry_price', 'exit_price', 'stop_loss', 'take_profit',
        'mark_price', 'liquidation_price',
        'profit_loss', 'profit_loss_percentage',
        'unrealized_pnl', 'unrealized_pnl_percentage',
        'max_loss_pct_reached', 'max_profit_pct_reached',
        'cut_loser_triggered', 'current_trailing_tier',
        'binance_order_id', 'binance_exit_order_id',
        'sl_order_id', 'tp_order_id', 'trailing_order_id',
        'signal_id', 'signal_timeframe', 'signal_confidence', 'signal_source',
        'error_message',
        'entry_time', 'exit_time', 'last_sync_time', 'created_at', 'updated_at',
    ]

    def _serialize_trade(self, trade):
        """Serialize a FuturesTrade to a flat dict suitable for export."""
        return {
            'id': trade.id,
            'symbol': trade.symbol,
            'direction': trade.direction,
            'status': trade.status,
            'leverage': trade.leverage,
            'margin_type': trade.margin_type,
            'quantity': self._to_float(trade.quantity),
            'position_size_usdt': self._to_float(trade.position_size_usdt),
            'entry_price': self._to_float(trade.entry_price),
            'exit_price': self._to_float(trade.exit_price),
            'stop_loss': self._to_float(trade.stop_loss),
            'take_profit': self._to_float(trade.take_profit),
            'mark_price': self._to_float(trade.mark_price),
            'liquidation_price': self._to_float(trade.liquidation_price),
            'profit_loss': self._to_float(trade.profit_loss),
            'profit_loss_percentage': self._to_float(trade.profit_loss_percentage),
            'unrealized_pnl': self._to_float(trade.unrealized_pnl),
            'unrealized_pnl_percentage': self._to_float(trade.unrealized_pnl_percentage),
            'max_loss_pct_reached': self._to_float(trade.max_loss_pct_reached),
            'max_profit_pct_reached': self._to_float(trade.max_profit_pct_reached),
            'cut_loser_triggered': trade.cut_loser_triggered,
            'current_trailing_tier': trade.current_trailing_tier,
            'binance_order_id': trade.binance_order_id,
            'binance_exit_order_id': trade.binance_exit_order_id,
            'sl_order_id': trade.sl_order_id,
            'tp_order_id': trade.tp_order_id,
            'trailing_order_id': trade.trailing_order_id,
            'signal_id': trade.signal.id if trade.signal else None,
            'signal_timeframe': trade.signal.timeframe if trade.signal else None,
            'signal_confidence': self._to_float(trade.signal.confidence) if trade.signal else None,
            'signal_source': trade.signal.source if trade.signal else None,
            'error_message': trade.error_message or '',
            'entry_time': self._to_iso(trade.entry_time),
            'exit_time': self._to_iso(trade.exit_time),
            'last_sync_time': self._to_iso(trade.last_sync_time),
            'created_at': self._to_iso(trade.created_at),
            'updated_at': self._to_iso(trade.updated_at),
        }

    @staticmethod
    def _to_float(value):
        """Convert a Decimal to float for JSON-safe export."""
        if value is None:
            return None
        if isinstance(value, Decimal):
            return float(value)
        return value

    @staticmethod
    def _to_iso(value):
        """Convert datetime to ISO 8601 string."""
        if value is None:
            return None
        return value.isoformat()

    @staticmethod
    def _filename(stem, ext):
        """Build a timestamped export filename."""
        ts = timezone.now().strftime('%Y%m%d_%H%M%S')
        return f'{stem}_{ts}.{ext}'

    def _export_queryset(self, queryset):
        """Return ordered queryset with related Signal preloaded."""
        return queryset.select_related('signal').order_by('-created_at')

    def _build_json_response(self, request, trades, filename):
        """Build a JSON HttpResponse for the given trades."""
        rows = [self._serialize_trade(t) for t in trades]
        payload = {
            'export_info': {
                'generated_at': timezone.now().isoformat(),
                'generated_by': request.user.username if request.user.is_authenticated else None,
                'total_trades': len(rows),
            },
            'trades': rows,
        }
        response = HttpResponse(
            json.dumps(payload, indent=2),
            content_type='application/json',
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    def _build_csv_response(self, trades, filename):
        """Build a CSV HttpResponse for the given trades."""
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        writer = csv.DictWriter(response, fieldnames=self.EXPORT_COLUMNS, extrasaction='ignore')
        writer.writeheader()
        for trade in trades:
            writer.writerow(self._serialize_trade(trade))
        return response

    def _build_excel_response(self, trades, filename):
        """Build an XLSX HttpResponse for the given trades."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            return None

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Futures Trades'

        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='1F2937', end_color='1F2937', fill_type='solid')
        sheet.append(self.EXPORT_COLUMNS)
        for col_idx in range(1, len(self.EXPORT_COLUMNS) + 1):
            cell = sheet.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill

        for trade in trades:
            row_data = self._serialize_trade(trade)
            sheet.append([row_data.get(col) for col in self.EXPORT_COLUMNS])

        for col_idx, column_name in enumerate(self.EXPORT_COLUMNS, start=1):
            sheet.column_dimensions[sheet.cell(row=1, column=col_idx).column_letter].width = max(
                12, min(len(column_name) + 4, 30)
            )
        sheet.freeze_panes = 'A2'

        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    @admin.action(description='Export selected trades to JSON')
    def export_selected_to_json(self, request, queryset):
        trades = self._export_queryset(queryset)
        return self._build_json_response(request, trades, self._filename('futures_trades', 'json'))

    @admin.action(description='Export selected trades to CSV')
    def export_selected_to_csv(self, request, queryset):
        trades = self._export_queryset(queryset)
        return self._build_csv_response(trades, self._filename('futures_trades', 'csv'))

    @admin.action(description='Export selected trades to Excel')
    def export_selected_to_excel(self, request, queryset):
        trades = self._export_queryset(queryset)
        response = self._build_excel_response(trades, self._filename('futures_trades', 'xlsx'))
        if response is None:
            self.message_user(
                request,
                'openpyxl is not installed; cannot export to Excel.',
                level=messages.ERROR,
            )
            return None
        return response

    def get_urls(self):
        from django.urls import path
        urls = super().get_urls()
        custom = [
            path(
                'export-all/json/',
                self.admin_site.admin_view(self.export_all_json_view),
                name='signals_futurestrade_export_json',
            ),
            path(
                'export-all/csv/',
                self.admin_site.admin_view(self.export_all_csv_view),
                name='signals_futurestrade_export_csv',
            ),
            path(
                'export-all/excel/',
                self.admin_site.admin_view(self.export_all_excel_view),
                name='signals_futurestrade_export_excel',
            ),
        ]
        return custom + urls

    def export_all_json_view(self, request):
        trades = self._export_queryset(FuturesTrade.objects.all())
        return self._build_json_response(request, trades, self._filename('futures_trades_all', 'json'))

    def export_all_csv_view(self, request):
        trades = self._export_queryset(FuturesTrade.objects.all())
        return self._build_csv_response(trades, self._filename('futures_trades_all', 'csv'))

    def export_all_excel_view(self, request):
        from django.shortcuts import redirect
        trades = self._export_queryset(FuturesTrade.objects.all())
        response = self._build_excel_response(trades, self._filename('futures_trades_all', 'xlsx'))
        if response is None:
            self.message_user(
                request,
                'openpyxl is not installed; cannot export to Excel.',
                level=messages.ERROR,
            )
            return redirect('..')
        return response


@admin.register(FuturesTradeLog)
class FuturesTradeLogAdmin(admin.ModelAdmin):
    """Admin interface for Futures Trade Logs - audit trail for all trade requests."""
    list_display = (
        'id', 'created_at', 'level_badge', 'action_badge',
        'symbol', 'direction_badge', 'priority_badge',
        'message_short', 'signal_link', 'trade_link',
    )
    list_filter = ('level', 'action', 'is_priority', 'force_execute', 'symbol', 'direction', 'created_at')
    search_fields = ('symbol', 'message', 'details')
    ordering = ('-created_at',)
    readonly_fields = (
        'signal', 'trade', 'action', 'level', 'symbol', 'direction',
        'is_priority', 'force_execute', 'message', 'details', 'created_at',
    )
    list_per_page = 50
    date_hierarchy = 'created_at'

    fieldsets = (
        ('Log Entry', {
            'fields': ('level', 'action', 'message', 'created_at')
        }),
        ('Trade Context', {
            'fields': ('signal', 'trade', 'symbol', 'direction', 'is_priority', 'force_execute')
        }),
        ('Details (JSON)', {
            'fields': ('details',),
            'classes': ('collapse',)
        }),
    )

    def has_add_permission(self, request):
        return False

    def has_change_permission(self, request, obj=None):
        return False

    def has_delete_permission(self, request, obj=None):
        return request.user.is_superuser

    @admin.display(description='Level')
    def level_badge(self, obj):
        colors = {
            'INFO': '#17a2b8',
            'WARNING': '#ffc107',
            'ERROR': '#dc3545',
            'SUCCESS': '#28a745',
        }
        text_color = '#000' if obj.level == 'WARNING' else '#fff'
        return format_html(
            '<span style="background-color: {}; color: {}; padding: 2px 8px; '
            'border-radius: 3px; font-size: 11px; font-weight: bold;">{}</span>',
            colors.get(obj.level, '#6c757d'), text_color, obj.level
        )

    @admin.display(description='Action')
    def action_badge(self, obj):
        colors = {
            'SIGNAL_RECEIVED': '#6f42c1',
            'CHECK_PASSED': '#20c997',
            'CHECK_FAILED': '#fd7e14',
            'TRADE_SUBMITTED': '#007bff',
            'TRADE_EXECUTED': '#28a745',
            'TRADE_FAILED': '#dc3545',
            'TRADE_CLOSED': '#6c757d',
            'ORDER_PLACED': '#17a2b8',
            'ORDER_FAILED': '#dc3545',
        }
        return format_html(
            '<span style="background-color: {}; color: white; padding: 2px 6px; '
            'border-radius: 3px; font-size: 10px;">{}</span>',
            colors.get(obj.action, '#6c757d'), obj.action.replace('_', ' ')
        )

    @admin.display(description='Dir')
    def direction_badge(self, obj):
        if not obj.direction:
            return '-'
        color = "#28a745" if obj.direction == "LONG" else "#dc3545"
        return format_html(
            '<span style="color: {}; font-weight: bold;">{}</span>',
            color, obj.direction
        )

    @admin.display(description='Priority')
    def priority_badge(self, obj):
        if obj.is_priority:
            return format_html(
                '<span style="background-color: #f59e0b; color: white; padding: 2px 6px; '
                'border-radius: 3px; font-size: 10px;">PRIORITY</span>'
            )
        return '-'

    @admin.display(description='Message')
    def message_short(self, obj):
        msg = obj.message[:80]
        if len(obj.message) > 80:
            msg += '...'
        return msg

    @admin.display(description='Signal')
    def signal_link(self, obj):
        if obj.signal_id:
            return format_html('<a href="/admin/signals/signal/{}/change/">#{}</a>', obj.signal_id, obj.signal_id)
        return '-'

    @admin.display(description='Trade')
    def trade_link(self, obj):
        if obj.trade_id:
            return format_html('<a href="/admin/signals/futurestrade/{}/change/">#{}</a>', obj.trade_id, obj.trade_id)
        return '-'

    def get_queryset(self, request):
        return super().get_queryset(request).select_related('signal', 'trade')


@admin.register(BlacklistedSymbol)
class BlacklistedSymbolAdmin(admin.ModelAdmin):
    """Admin interface for blacklisted symbols."""
    list_display = (
        'symbol',
        'reason_badge',
        'active_badge',
        'blacklisted_at',
        'blacklisted_until',
        'notes_short',
    )
    list_filter = ('active', 'reason', 'blacklisted_at')
    search_fields = ('symbol', 'notes')
    readonly_fields = ('created_at', 'updated_at')
    ordering = ('-blacklisted_at',)
    date_hierarchy = 'blacklisted_at'

    fieldsets = (
        ('Symbol Information', {
            'fields': ('symbol', 'active')
        }),
        ('Blacklist Details', {
            'fields': ('reason', 'notes', 'blacklisted_at', 'blacklisted_until')
        }),
        ('Metadata', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )

    def reason_badge(self, obj):
        """Display reason with color badge."""
        color_map = {
            'HIGH_VOLATILITY': '#dc3545',  # Red
            'LOW_LIQUIDITY': '#ffc107',    # Yellow
            'POOR_PERFORMANCE': '#fd7e14', # Orange
            'DELISTED': '#6c757d',         # Gray
            'TEMPORARY': '#17a2b8',        # Cyan
            'MANUAL': '#6f42c1',           # Purple
            'OTHER': '#6c757d',            # Gray
        }
        color = color_map.get(obj.reason, '#6c757d')
        return mark_safe(
            f'<span style="background-color: {color}; color: white; padding: 3px 8px; '
            f'border-radius: 3px; font-size: 11px;">{obj.get_reason_display()}</span>'
        )
    reason_badge.short_description = 'Reason'

    def active_badge(self, obj):
        """Display active status with badge."""
        if obj.is_expired():
            color = '#6c757d'
            text = 'EXPIRED'
        elif obj.active:
            color = '#dc3545'
            text = 'ACTIVE'
        else:
            color = '#28a745'
            text = 'INACTIVE'
        return mark_safe(
            f'<span style="background-color: {color}; color: white; padding: 3px 8px; '
            f'border-radius: 3px; font-weight: bold; font-size: 11px;">{text}</span>'
        )
    active_badge.short_description = 'Status'

    def notes_short(self, obj):
        """Display shortened notes."""
        if not obj.notes:
            return '-'
        if len(obj.notes) > 50:
            return f"{obj.notes[:50]}..."
        return obj.notes
    notes_short.short_description = 'Notes'

    actions = ['activate_blacklist', 'deactivate_blacklist', 'remove_expiration']

    def activate_blacklist(self, request, queryset):
        """Activate selected blacklist entries."""
        updated = queryset.update(active=True)
        self.message_user(
            request,
            f'{updated} blacklist entries were successfully activated.',
            messages.SUCCESS
        )
    activate_blacklist.short_description = 'Activate selected blacklists'

    def deactivate_blacklist(self, request, queryset):
        """Deactivate selected blacklist entries."""
        updated = queryset.update(active=False)
        self.message_user(
            request,
            f'{updated} blacklist entries were successfully deactivated.',
            messages.SUCCESS
        )
    deactivate_blacklist.short_description = 'Deactivate selected blacklists'

    def remove_expiration(self, request, queryset):
        """Remove expiration date (make permanent)."""
        updated = queryset.update(blacklisted_until=None)
        self.message_user(
            request,
            f'{updated} blacklist entries are now permanent (no expiration).',
            messages.SUCCESS
        )
    remove_expiration.short_description = 'Make permanent (remove expiration)'


@admin.register(PushSubscription)
class PushSubscriptionAdmin(admin.ModelAdmin):
    """Admin interface for push notification subscriptions."""
    list_display = ('user', 'device_name', 'is_active', 'created_at', 'updated_at')
    list_filter = ('is_active', 'created_at')
    search_fields = ('user__username', 'device_name', 'fcm_token')
    readonly_fields = ('created_at', 'updated_at')
    ordering = ('-created_at',)
    actions = ['deactivate_subscriptions', 'activate_subscriptions', 'send_test_notification', 'send_test_signal', 'send_test_session']

    def deactivate_subscriptions(self, request, queryset):
        """Deactivate selected push subscriptions."""
        updated = queryset.update(is_active=False)
        self.message_user(request, f'{updated} subscriptions deactivated.', messages.SUCCESS)
    deactivate_subscriptions.short_description = 'Deactivate selected subscriptions'

    def activate_subscriptions(self, request, queryset):
        """Activate selected push subscriptions."""
        updated = queryset.update(is_active=True)
        self.message_user(request, f'{updated} subscriptions activated.', messages.SUCCESS)
    activate_subscriptions.short_description = 'Activate selected subscriptions'

    def send_test_notification(self, request, queryset):
        """Send a test push notification to selected subscribers."""
        from signals.services.push_notification import _send_multicast
        tokens = list(queryset.filter(is_active=True).values_list('fcm_token', flat=True))
        if not tokens:
            self.message_user(request, 'No active subscriptions selected.', messages.WARNING)
            return
        result = _send_multicast(tokens, 'RevX Test', 'Push notifications are working!', {'type': 'TEST'})
        self.message_user(request, f'Test notification: {result["sent"]}/{result["total"]} sent.', messages.SUCCESS)
    send_test_notification.short_description = 'Send test notification to selected'

    def send_test_signal(self, request, queryset):
        """Send a simulated priority signal notification to selected subscribers."""
        from signals.services.push_notification import _send_multicast
        tokens = list(queryset.filter(is_active=True).values_list('fcm_token', flat=True))
        if not tokens:
            self.message_user(request, 'No active subscriptions selected.', messages.WARNING)
            return
        result = _send_multicast(
            tokens,
            '\U0001F7E2 LONG BTCUSDT [PRIORITY]',
            'Entry: $84,500 | SL: $83,200 | TP: $87,000 | Conf: 85%',
            {'type': 'NEW_SIGNAL', 'symbol': 'BTCUSDT', 'direction': 'LONG', 'is_priority': 'true'},
        )
        self.message_user(request, f'Signal notification: {result["sent"]}/{result["total"]} sent.', messages.SUCCESS)
    send_test_signal.short_description = 'Send test SIGNAL notification to selected'

    def send_test_session(self, request, queryset):
        """Send a simulated session activation notification to selected subscribers."""
        from signals.services.push_notification import _send_multicast
        tokens = list(queryset.filter(is_active=True).values_list('fcm_token', flat=True))
        if not tokens:
            self.message_user(request, 'No active subscriptions selected.', messages.WARNING)
            return
        result = _send_multicast(
            tokens,
            'Trading Session Active - GW1',
            'Started at 10:15 AM NPT | Ends at 11:30 NPT | Priority signals will auto-trade',
            {'type': 'SESSION_ACTIVE', 'session_name': 'GW1'},
        )
        self.message_user(request, f'Session notification: {result["sent"]}/{result["total"]} sent.', messages.SUCCESS)
    send_test_session.short_description = 'Send test SESSION notification to selected'


@admin.register(NotificationLog)
class NotificationLogAdmin(admin.ModelAdmin):
    """Admin interface for push notification audit logs with send functionality."""
    list_display = ('title', 'status_badge', 'user', 'tokens_targeted', 'tokens_succeeded', 'signal', 'created_at')
    list_filter = ('status', 'created_at')
    search_fields = ('title', 'body', 'user__username')
    readonly_fields = ('user', 'title', 'body', 'data', 'status', 'error_message', 'signal', 'tokens_targeted', 'tokens_succeeded', 'created_at')
    ordering = ('-created_at',)
    date_hierarchy = 'created_at'

    def has_add_permission(self, request):
        return True

    def add_view(self, request, form_url='', extra_context=None):
        """Override add view to show send notification form."""
        from django.shortcuts import render, redirect

        if request.method == 'POST':
            title = request.POST.get('title', '').strip()
            body = request.POST.get('body', '').strip()
            target = request.POST.get('target', 'all')

            if not title or not body:
                messages.error(request, 'Title and body are required.')
                return render(request, 'admin/send_push_notification.html', {
                    'title': 'Send Push Notification',
                    'opts': self.model._meta,
                    'has_view_permission': True,
                    'form_title': title,
                    'form_body': body,
                    'form_target': target,
                    'subscriber_count': PushSubscription.objects.filter(is_active=True).count(),
                })

            from signals.services.push_notification import broadcast, send_to_user

            if target == 'all':
                result = broadcast(title, body, data={'type': 'ADMIN_BROADCAST'})
            else:
                from django.contrib.auth import get_user_model
                User = get_user_model()
                try:
                    user = User.objects.get(pk=int(target))
                    result = send_to_user(user, title, body, data={'type': 'ADMIN_MESSAGE'})
                except User.DoesNotExist:
                    messages.error(request, f'User ID {target} not found.')
                    return redirect('..')

            if result['sent'] > 0:
                messages.success(request, f'Notification sent: {result["sent"]}/{result["total"]} delivered.')
            elif result['total'] == 0:
                messages.warning(request, 'No active subscribers found.')
            else:
                messages.error(request, f'Send failed: {result["error"]}')

            return redirect('..')

        from django.contrib.auth import get_user_model
        User = get_user_model()
        users_with_subs = (
            PushSubscription.objects.filter(is_active=True)
            .values_list('user__id', 'user__username')
            .distinct()
        )

        context = {
            'title': 'Send Push Notification',
            'opts': self.model._meta,
            'has_view_permission': True,
            'form_title': '',
            'form_body': '',
            'form_target': 'all',
            'subscriber_count': PushSubscription.objects.filter(is_active=True).count(),
            'users_with_subs': list(users_with_subs),
        }
        return render(request, 'admin/send_push_notification.html', context)

    def status_badge(self, obj):
        """Display status with color badge."""
        color_map = {'SENT': '#28a745', 'FAILED': '#dc3545', 'PARTIAL': '#ffc107'}
        color = color_map.get(obj.status, '#6c757d')
        return mark_safe(
            f'<span style="background-color: {color}; color: white; padding: 3px 8px; '
            f'border-radius: 3px; font-size: 11px;">{obj.status}</span>'
        )
    status_badge.short_description = 'Status'


class BacktestTradeInline(admin.TabularInline):
    model = BacktestTrade
    extra = 0
    readonly_fields = (
        'symbol', 'direction', 'entry_price', 'exit_price',
        'stop_loss', 'take_profit', 'profit_loss', 'profit_loss_percentage',
        'status', 'opened_at', 'closed_at', 'duration_hours',
    )
    fields = readonly_fields
    can_delete = False
    show_change_link = False

    def has_add_permission(self, request, obj=None):
        return False


@admin.register(BacktestRun)
class BacktestRunAdmin(admin.ModelAdmin):
    list_display = (
        'id', 'name', 'status_badge', 'symbols_display', 'timeframe',
        'total_trades', 'win_rate_display', 'roi_display', 'created_at',
    )
    list_filter = ('status', 'timeframe', 'created_at')
    search_fields = ('name',)
    readonly_fields = (
        'status', 'started_at', 'completed_at', 'error_message',
        'total_trades', 'winning_trades', 'losing_trades', 'win_rate',
        'total_profit_loss', 'roi', 'max_drawdown', 'sharpe_ratio',
        'profit_factor', 'progress_pct', 'created_at', 'updated_at',
    )
    inlines = [BacktestTradeInline]
    date_hierarchy = 'created_at'
    list_per_page = 25
    ordering = ('-created_at',)

    fieldsets = (
        ('Configuration', {
            'fields': ('name', 'user', 'symbols', 'timeframe', 'start_date', 'end_date',
                       'initial_capital', 'position_size', 'strategy_params'),
        }),
        ('Execution', {
            'fields': ('status', 'progress_pct', 'started_at', 'completed_at', 'error_message'),
        }),
        ('Results', {
            'fields': ('total_trades', 'winning_trades', 'losing_trades', 'win_rate',
                       'total_profit_loss', 'roi', 'max_drawdown', 'sharpe_ratio', 'profit_factor'),
        }),
    )

    def status_badge(self, obj):
        colors = {'COMPLETED': '#28a745', 'FAILED': '#dc3545', 'RUNNING': '#007bff', 'PENDING': '#6c757d'}
        color = colors.get(obj.status, '#6c757d')
        return mark_safe(
            f'<span style="background:{color};color:white;padding:2px 8px;border-radius:3px;font-size:11px;">'
            f'{obj.status}</span>'
        )
    status_badge.short_description = 'Status'

    def symbols_display(self, obj):
        return ', '.join(obj.symbols) if obj.symbols else '-'
    symbols_display.short_description = 'Symbols'

    def win_rate_display(self, obj):
        if obj.win_rate:
            return f'{obj.win_rate:.1f}%'
        return '-'
    win_rate_display.short_description = 'Win Rate'

    def roi_display(self, obj):
        if obj.roi is None:
            return '-'
        color = '#28a745' if obj.roi >= 0 else '#dc3545'
        return mark_safe(f'<span style="color:{color};font-weight:bold;">{obj.roi:+.2f}%</span>')
    roi_display.short_description = 'ROI'


@admin.register(BacktestTrade)
class BacktestTradeAdmin(admin.ModelAdmin):
    list_display = (
        'id', 'backtest_run', 'symbol', 'direction_badge', 'entry_price',
        'exit_price', 'pnl_display', 'status', 'duration_hours',
    )
    list_filter = ('direction', 'status', 'symbol')
    readonly_fields = (
        'backtest_run', 'symbol', 'direction', 'entry_price', 'exit_price',
        'stop_loss', 'take_profit', 'position_size', 'quantity',
        'profit_loss', 'profit_loss_percentage', 'status',
        'opened_at', 'closed_at', 'duration_hours',
    )
    list_per_page = 50

    def direction_badge(self, obj):
        color = '#28a745' if obj.direction == 'LONG' else '#dc3545'
        return mark_safe(
            f'<span style="background:{color};color:white;padding:2px 6px;border-radius:3px;font-size:11px;">'
            f'{obj.direction}</span>'
        )
    direction_badge.short_description = 'Direction'

    def pnl_display(self, obj):
        if obj.profit_loss is None:
            return '-'
        color = '#28a745' if obj.profit_loss >= 0 else '#dc3545'
        return mark_safe(f'<span style="color:{color};font-weight:bold;">${obj.profit_loss:+.2f}</span>')
    pnl_display.short_description = 'P/L'


@admin.register(BacktestMetric)
class BacktestMetricAdmin(admin.ModelAdmin):
    list_display = ('id', 'backtest_run', 'timestamp', 'equity', 'total_pnl', 'total_trades')
    list_filter = ('backtest_run',)
    readonly_fields = (
        'backtest_run', 'timestamp', 'equity', 'cash', 'open_positions_value',
        'total_pnl', 'unrealized_pnl', 'realized_pnl',
        'total_trades', 'winning_trades', 'losing_trades', 'open_trades',
    )
    list_per_page = 50


@admin.register(BalanceRebalanceLog)
class BalanceRebalanceLogAdmin(admin.ModelAdmin):
    list_display = (
        'created_at', 'applied', 'balance', 'per_trade_amount',
        'max_concurrent_trades', 'backup_reserve',
        'previous_trade_amount', 'reason',
    )
    list_filter = ('applied', 'created_at')
    readonly_fields = (
        'created_at', 'balance', 'per_trade_amount', 'max_concurrent_trades',
        'backup_reserve', 'previous_trade_amount',
        'previous_max_concurrent_trades', 'applied', 'reason',
    )
    ordering = ('-created_at',)
    list_per_page = 50

    def has_add_permission(self, request):
        return False

    def has_change_permission(self, request, obj=None):
        return False


from .models.daytrade import (
    DayTradeSignal,
    DayTradePaperTrade,
    DayTradeTradeExit,
    DayTradePaperAccount,
    DayTradeStrategyConfig,
    DayTradeSession,
)


@admin.register(DayTradeSession)
class DayTradeSessionAdmin(admin.ModelAdmin):
    """Auto-discovered favourable day-trade windows (from the session optimizer)."""

    list_display = (
        'name', 'session_type', 'start_hour', 'end_hour', 'active_days',
        'win_rate', 'total_trades_analyzed', 'is_active', 'auto_generated',
        'last_optimized_at',
    )
    list_filter = ('is_active', 'session_type', 'auto_generated')
    search_fields = ('name',)
    ordering = ('start_hour', 'name')
    readonly_fields = ('created_at', 'updated_at', 'last_optimized_at')
    list_editable = ('is_active',)


def _export_field_names(queryset):
    """Concrete field names of the queryset's model, in declared order."""
    return [field.name for field in queryset.model._meta.fields]


@admin.action(description='Export selected as JSON')
def export_as_json(modeladmin, request, queryset):
    """Download the selected rows as a JSON file."""
    fields = _export_field_names(queryset)
    payload = json.dumps(list(queryset.values(*fields)), default=str, indent=2)
    response = HttpResponse(payload, content_type='application/json')
    response['Content-Disposition'] = 'attachment; filename="daytrade_export.json"'
    return response


@admin.action(description='Export selected as CSV')
def export_as_csv(modeladmin, request, queryset):
    """Download the selected rows as a CSV file."""
    fields = _export_field_names(queryset)
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="daytrade_export.csv"'
    writer = csv.writer(response)
    writer.writerow(fields)
    for row in queryset.values_list(*fields):
        writer.writerow(row)
    return response


@admin.action(description='Export selected as Excel (xlsx)')
def export_as_xlsx(modeladmin, request, queryset):
    """Download the selected rows as an Excel workbook."""
    from openpyxl import Workbook

    fields = _export_field_names(queryset)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = queryset.model._meta.verbose_name_plural[:31]
    sheet.append(fields)
    for row in queryset.values_list(*fields):
        sheet.append(['' if value is None else str(value) for value in row])
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="daytrade_export.xlsx"'
    workbook.save(response)
    return response


DAYTRADE_EXPORT_ACTIONS = [export_as_json, export_as_csv, export_as_xlsx]


@admin.register(DayTradeSignal)
class DayTradeSignalAdmin(admin.ModelAdmin):
    """Day-trade signals emitted by the 15m Market Structure engine."""

    list_display = (
        'symbol', 'direction', 'entry_timeframe', 'entry', 'stop_loss',
        'tp1', 'tp2', 'confidence', 'score', 'status', 'created_at',
    )
    list_filter = ('status', 'direction', 'entry_timeframe', 'symbol')
    search_fields = ('symbol',)
    ordering = ('-created_at',)
    readonly_fields = ('created_at', 'updated_at')
    list_per_page = 50
    actions = DAYTRADE_EXPORT_ACTIONS


@admin.register(DayTradePaperTrade)
class DayTradePaperTradeAdmin(admin.ModelAdmin):
    """Day-trade paper positions with scale-out and trailing-stop state."""

    list_display = (
        'symbol', 'direction', 'status', 'confidence', 'entry_price', 'remaining_quantity',
        'stop_loss', 'trailing_stop', 'tp1_filled', 'tp2_filled',
        'profit_loss', 'entry_time',
    )
    list_filter = ('status', 'direction', 'symbol', 'tp1_filled', 'tp2_filled')
    search_fields = ('symbol',)
    ordering = ('-created_at',)
    readonly_fields = ('created_at', 'updated_at')
    list_per_page = 50
    actions = DAYTRADE_EXPORT_ACTIONS


@admin.register(DayTradeTradeExit)
class DayTradeTradeExitAdmin(admin.ModelAdmin):
    """Individual scale-out legs (TP1/TP2/trail/SL) of a day-trade."""

    list_display = ('trade', 'exit_type', 'price', 'quantity', 'pnl', 'exit_time')
    list_filter = ('exit_type',)
    search_fields = ('trade__symbol',)
    ordering = ('-exit_time',)
    list_per_page = 50


@admin.register(DayTradePaperAccount)
class DayTradePaperAccountAdmin(admin.ModelAdmin):
    """Virtual account tracking day-trade bot performance."""

    list_display = (
        '__str__', 'balance', 'equity', 'total_pnl', 'realized_pnl',
        'win_rate', 'total_trades', 'winning_trades', 'losing_trades',
        'updated_at',
    )
    readonly_fields = ('created_at', 'updated_at')


@admin.register(DayTradeStrategyConfig)
class DayTradeStrategyConfigAdmin(admin.ModelAdmin):
    """Tunable parameters that drive the 15m Market Structure engine."""

    list_display = (
        'name', 'is_active', 'entry_timeframe', 'trend_timeframe',
        'min_confidence', 'margin_per_trade', 'leverage',
        'sl_atr_mult', 'tp1_atr_mult', 'tp2_atr_mult',
        'min_score', 'updated_at',
    )
    list_filter = ('is_active',)
    list_editable = ('is_active',)
    readonly_fields = ('created_at', 'updated_at')

    fieldsets = (
        ('Identity & Universe', {
            'fields': ('name', 'is_active', 'symbols', 'universe_top_n', 'entry_timeframe', 'trend_timeframe'),
        }),
        ('Trend Filter (1H)', {
            'fields': ('trend_ema_fast', 'trend_ema_slow'),
        }),
        ('Market Structure', {
            'fields': ('pivot_lookback',),
        }),
        ('Pullback Zone', {
            'fields': ('pullback_ema_fast', 'pullback_ema_slow', 'use_vwap', 'vwap_anchor'),
        }),
        ('Momentum', {
            'fields': ('rsi_period', 'rsi_threshold', 'macd_fast', 'macd_slow', 'macd_signal'),
        }),
        ('Volume & Trend Strength', {
            'fields': ('volume_multiplier', 'volume_avg_period', 'adx_min', 'adx_period'),
        }),
        ('SL/TP (v1-style fixed percentage)', {
            'fields': ('sl_percentage', 'tp_percentage'),
            'description': 'Single stop and take-profit as percent from entry (active exit model)',
        }),
        ('Risk: ATR (legacy/scoring only)', {
            'classes': ('collapse',),
            'fields': (
                'atr_period', 'sl_atr_mult',
                'tp1_atr_mult', 'tp1_close_pct',
                'tp2_atr_mult', 'tp2_close_pct',
                'runner_pct', 'trail_atr_mult',
                'risk_per_trade_pct',
            ),
        }),
        ('Position Sizing (paper)', {
            'fields': ('margin_per_trade', 'leverage'),
            'description': 'Each paper trade uses this fixed margin and leverage',
        }),
        ('Confirmations', {
            'fields': ('enable_liquidity_sweep',),
        }),
        ('Scoring Weights', {
            'fields': (
                'weight_trend', 'weight_structure', 'weight_volume',
                'weight_pullback', 'weight_macd', 'weight_rsi', 'weight_atr',
                'min_score', 'min_confidence',
            ),
        }),
        ('V3 Trend-strength gate (validated: ON)', {
            'fields': (
                'trend_filter_enabled', 'trend_require_price_above_ema50',
                'trend_min_ema_gap_pct', 'trend_min_slope_pct',
                'trend_slope_lookback', 'trend_require_adx_rising',
            ),
            'description': 'Validated combo: enabled + price-above-EMA50 + EMA gap >= 0.5%. Slope/ADX-rising left off.',
        }),
        ('V3 Market-regime gate (validated: ON, ATR percentile only)', {
            'fields': (
                'regime_filter_enabled', 'regime_atr_percentile_min',
                'regime_atr_percentile_period', 'regime_min_adx',
                'regime_max_choppiness', 'regime_choppiness_period',
                'regime_min_bbw_pct', 'regime_bb_period', 'regime_bb_std',
            ),
            'description': 'Validated: enabled + ATR percentile >= 30. ADX/choppiness/BBW left off (did not generalise).',
        }),
        ('V3 Structure quality (validated: OFF)', {
            'classes': ('collapse',),
            'fields': (
                'structure_quality_enabled', 'structure_min_swing_atr',
                'weight_structure_bonus', 'require_bos', 'block_on_choch',
            ),
            'description': 'Not robust in walk-forward testing; left off.',
        }),
        ('Timestamps', {
            'fields': ('created_at', 'updated_at'),
        }),
    )



from signals.models.swing import SwingStrategyConfig, SwingPaperTrade


@admin.register(SwingStrategyConfig)
class SwingStrategyConfigAdmin(admin.ModelAdmin):
    """Admin for the 4h swing engine config (paper harness)."""
    list_display = (
        'name', 'enabled', 'entry_timeframe', 'trend_timeframe',
        'adx_min', 'breakout_lookback', 'sl_atr_mult', 'tp_atr_mult',
        'margin_per_trade', 'leverage', 'updated_at',
    )


@admin.register(SwingPaperTrade)
class SwingPaperTradeAdmin(admin.ModelAdmin):
    """Admin for 4h swing paper trades (read-only)."""
    list_display = (
        'symbol', 'direction', 'status', 'entry_price', 'exit_price',
        'profit_loss', 'fees_paid', 'entry_time', 'exit_time',
    )
    list_filter = ('status', 'direction', 'symbol')
    search_fields = ('symbol',)
    readonly_fields = [f.name for f in SwingPaperTrade._meta.fields]
